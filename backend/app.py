from flask import Flask, request, jsonify, send_file, g, Response
from flask_cors import CORS
import os
import requests
import csv
import io
from datetime import datetime, timedelta
import json
import sys
import logging
import base64
import time
import hashlib
from flask_socketio import SocketIO, emit
import asyncio
from typing import List, Dict
from contextlib import contextmanager
from auth import (
    auth_bp,
    token_required,
    _effective_user_ids_for_data_scope,
    sync_account_verify_to_db,
    build_validate_api_payload,
)
import uuid
from database.db import get_db_cursor, init_db, connection_pool
from Config import get_account  # Fix import casing
import urllib.parse
try:
    import redis  # type: ignore[reportMissingImports]
except Exception:  # pragma: no cover - optional dependency
    redis = None

def format_timestamp(timestamp_str):
    """
    Strip milliseconds from timestamps.
    e.g. 2025-10-13 23:00:00.000 -> 2025-10-13 23:00:00
    """
    try:
        if not timestamp_str:
            return ''

        ts = timestamp_str.strip()
        if not ts:
            return ''

        # Already in target format
        if len(ts) == 19 and ts[4] == '-' and ts[7] == '-' and ts[10] == ' ' and ts[13] == ':' and ts[16] == ':':
            return ts

        # Common timestamp formats; align with AutoPipe where possible
        formats = [
            '%Y-%m-%d %H:%M:%S.%f',
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%dT%H:%M:%S.%f',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%dT%H:%M:%SZ',
            '%Y-%m-%dT%H:%M:%S.%fZ',
        ]
        for fmt in formats:
            try:
                return datetime.strptime(ts, fmt).strftime('%Y-%m-%d %H:%M:%S')
            except ValueError:
                continue

        # Keep original value when format is unknown
        if '.' in ts and ' ' in ts:
            return ts.split('.')[0]
        return ts
    except Exception:
        return timestamp_str


def _normalize_header_name(header):
    normalized = (header or '').strip().lower().replace('\ufeff', '')
    normalized = normalized.replace('-', '_').replace(' ', '_')
    while '__' in normalized:
        normalized = normalized.replace('__', '_')
    return normalized


def _get_query_param_case_insensitive(query_params, *keys):
    for target_key in keys:
        value = query_params.get(target_key)
        if value:
            return value[0]
        for actual_key, actual_values in query_params.items():
            if actual_key.lower() == target_key.lower() and actual_values:
                return actual_values[0]
    return ''


def _parse_postback_url_data(postback_url):
    result = {}
    if not postback_url:
        return result

    try:
        parsed_url = urllib.parse.urlparse(postback_url)
        query_params = urllib.parse.parse_qs(parsed_url.query)
    except Exception:
        return result

    def _decoded(*keys):
        raw = _get_query_param_case_insensitive(query_params, *keys)
        if not raw:
            return ''
        return urllib.parse.unquote(raw).strip()

    click_time = _decoded('click_timestamp', 'click_time', 'attributed_touch_time')
    if click_time:
        result['attributed_touch_time'] = format_timestamp(click_time)

    install_time = _decoded('install_timestamp', 'install_time')
    if install_time:
        result['install_time'] = format_timestamp(install_time)

    download_time = _decoded('download_timestamp')
    if download_time and not result.get('install_time'):
        result['install_time'] = format_timestamp(download_time)

    event_time = _decoded('timestamp', 'time', 'event_time')
    if event_time:
        result['event_time'] = format_timestamp(event_time)

    field_mapping = {
        'appsflyer_id': ('appsflyer_id',),
        'advertising_id': ('gaid', 'advertising_id'),
        'idfa': ('idfa',),
        'idfv': ('idfv',),
        'customer_user_id': ('user_id', 'customer_user_id'),
        'ip': ('ip',),
        'language': ('lang', 'language'),
        'app_version': ('app_version',),
        'os_version': ('os_version',),
        'bundle_id': ('bundle_id',),
        'match_type': ('match_type',),
    }
    for target_field, aliases in field_mapping.items():
        value = _decoded(*aliases)
        if value:
            result[target_field] = value

    # Support both country and country_code params
    country = _decoded('country_code', 'country')
    if country:
        result['country_code'] = country

    # Fall back to IDFA for advertising_id when missing
    if not result.get('advertising_id') and result.get('idfa'):
        result['advertising_id'] = result['idfa']

    return result

def process_postback_urls_in_csv(csv_content):
    """
    Fill missing fields from Postback URL per CSV row.
    AutoPipe rules: iOS rows only, fill blanks only, case-insensitive params.
    """
    try:
        csv_reader = csv.reader(io.StringIO(csv_content))
        rows = list(csv_reader)
        if len(rows) < 2:
            return csv_content

        headers = rows[0]

        normalized_headers = [_normalize_header_name(header) for header in headers]
        index_map = {}
        for i, normalized_name in enumerate(normalized_headers):
            if normalized_name and normalized_name not in index_map:
                index_map[normalized_name] = i

        def _find_index(*names):
            for name in names:
                normalized = _normalize_header_name(name)
                if normalized in index_map:
                    return index_map[normalized]
            return None

        platform_index = _find_index('platform')
        postback_url_index = _find_index('postback_url', 'postback url')
        app_id_index = _find_index('app_id', 'app id', 'app_id_ios')

        target_field_indexes = {
            'attributed_touch_time': _find_index('attributed_touch_time', 'attributed touch time'),
            'install_time': _find_index('install_time', 'install time'),
            'event_time': _find_index('event_time', 'event time', 'timestamp'),
            'appsflyer_id': _find_index('appsflyer_id', 'appsflyer id'),
            'advertising_id': _find_index('advertising_id', 'advertising id', 'gaid'),
            'idfa': _find_index('idfa'),
            'idfv': _find_index('idfv'),
            'customer_user_id': _find_index('customer_user_id', 'customer user id', 'user_id'),
            'ip': _find_index('ip'),
            'language': _find_index('language', 'lang'),
            'app_version': _find_index('app_version', 'app version'),
            'os_version': _find_index('os_version', 'os version'),
            'bundle_id': _find_index('bundle_id', 'bundle id'),
            'match_type': _find_index('match_type', 'match type'),
            'country_code': _find_index('country_code', 'country code', 'country'),
        }

        if postback_url_index is None:
            logger.info("未找到Postback URL字段，跳过处理")
            return csv_content

        def _ensure_len(row):
            while len(row) < len(headers):
                row.append('')

        def _is_ios_row(row):
            if platform_index is not None and platform_index < len(row):
                if (row[platform_index] or '').strip().lower() == 'ios':
                    return True
            if app_id_index is not None and app_id_index < len(row):
                app_id = (row[app_id_index] or '').strip().lower()
                if app_id and (app_id.startswith('id') or app_id.isdigit()):
                    return True
            return False

        enhanced_rows = 0
        ios_rows = 0
        processed_rows = 0

        for i in range(1, len(rows)):
            row = rows[i]
            if not row:
                continue

            _ensure_len(row)

            if not _is_ios_row(row):
                continue

            ios_rows += 1
            postback_url = (row[postback_url_index] or '').strip()
            if not postback_url:
                continue

            extracted = _parse_postback_url_data(postback_url)
            if i == 1 and extracted:
                logger.debug(f"第一行提取字段: {list(extracted.keys())}")
            if not extracted:
                continue

            changed = False
            install_idx = target_field_indexes.get('install_time')
            touch_idx = target_field_indexes.get('attributed_touch_time')
            if install_idx is not None and touch_idx is not None and extracted.get('attributed_touch_time'):
                if not (row[install_idx] or '').strip():
                    row[install_idx] = extracted['attributed_touch_time']
                    changed = True

            for field_name, value in extracted.items():
                field_index = target_field_indexes.get(field_name)
                if field_index is None or not value:
                    continue
                if not (row[field_index] or '').strip():
                    row[field_index] = value
                    changed = True

            if changed:
                enhanced_rows += 1
            processed_rows += 1

        output = io.StringIO()
        csv_writer = csv.writer(output)
        csv_writer.writerows(rows)
        result = output.getvalue()

        logger.info(
            f"Postback URL处理完成：总行数 {len(rows)-1}，iOS行 {ios_rows}，"
            f"含Postback URL行 {processed_rows}，增强成功 {enhanced_rows} 行"
        )

        return result
    except Exception as e:
        logger.error(f"处理CSV中的Postback URL失败: {str(e)}")
        return csv_content
from flask_jwt_extended import jwt_required, get_jwt_identity, JWTManager  # type: ignore[reportMissingImports]
import re
import zipfile
import shutil
from werkzeug.exceptions import HTTPException
import jwt  # type: ignore[reportMissingImports]
from auth import JWT_SECRET_KEY
import urllib.parse
from io import BytesIO
import openpyxl  # type: ignore[reportMissingModuleSource]
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter  # type: ignore[reportMissingModuleSource]
import urllib3
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import ssl

# Load environment variables
from dotenv import load_dotenv  # type: ignore[reportMissingImports]
load_dotenv()

# Configure logging
logging.basicConfig(
    level=os.getenv('LOG_LEVEL', 'INFO'),  # Default log level: INFO
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Configure global SSL settings
def configure_global_ssl():
    """Configure global SSL settings."""
    try:
        # Configure urllib3 SSL
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        # Set default SSL context
        ssl_context = ssl.create_default_context()
        ssl_context.minimum_version = ssl.TLSVersion.TLSv1_2
        ssl_context.maximum_version = ssl.TLSVersion.TLSv1_3
        
        # Set default requests SSL config
        import requests
        
        logger.info('全局SSL配置已设置')
    except Exception as e:
        logger.warning(f'全局SSL配置失败: {str(e)}')

# Configure SSL at app startup
configure_global_ssl()

# Log env vars (debug)
logger.info(f"LOG_LEVEL: {os.getenv('LOG_LEVEL', 'INFO')}")
logger.info(f"DB_HOST: {os.getenv('DB_HOST', 'localhost')}")
logger.info(f"DB_NAME: {os.getenv('DB_NAME', 'appsflyer_rawdata')}")
logger.info(f"IS_LOCAL: {os.getenv('IS_LOCAL', 'false')}")

# Add parent dir to sys.path
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

app = Flask(__name__)

# Initialize DB config
init_db(app)

# JWT config
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', 'change-me-in-production')
app.config['JWT_TOKEN_LOCATION'] = ['headers']
app.config['JWT_HEADER_NAME'] = 'Authorization'
app.config['JWT_HEADER_TYPE'] = 'Bearer'
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=12)
app.config['JWT_ERROR_MESSAGE_KEY'] = 'msg'
app.config['JWT_ALGORITHM'] = 'HS256'
app.config['JWT_DECODE_AUDIENCE'] = None
app.config['JWT_DECODE_ISSUER'] = None
app.config['JWT_DECODE_LEEWAY'] = 0
app.config['JWT_VERIFY_EXPIRATION'] = True
app.config['JWT_VERIFY_CLAIMS'] = ['exp', 'iat', 'nbf', 'sub']

# Init JWT (do not name variable jwt; avoids shadowing PyJWT module used by jwt.decode / ExpiredSignatureError)
jwt_manager = JWTManager(app)

HOME_CACHE_GENERATION_KEY = "autopipe:home:data:generation"
HOME_CACHE_PREFIX = "autopipe:home:v1"
HOME_CACHE_TTL_QUERY_RESULTS = int(os.getenv("HOME_CACHE_TTL_QUERY_RESULTS_SECONDS", "20"))
HOME_CACHE_TTL_PREVIEW = int(os.getenv("HOME_CACHE_TTL_PREVIEW_SECONDS", "30"))
HOME_CACHE_TTL_APP_NAME = int(os.getenv("HOME_CACHE_TTL_APP_NAME_SECONDS", "300"))
APPS_FINDER_CACHE_GENERATION_KEY = "autopipe:appsfinder:data:generation"
APPS_FINDER_CACHE_PREFIX = "autopipe:appsfinder:v1"
APPS_FINDER_CACHE_TTL_LIST = int(os.getenv("APPS_FINDER_CACHE_TTL_LIST_SECONDS", "60"))
APPS_FINDER_CACHE_TTL_OPTIONS = int(os.getenv("APPS_FINDER_CACHE_TTL_OPTIONS_SECONDS", "90"))
APPS_FINDER_CACHE_TTL_DETAIL = int(os.getenv("APPS_FINDER_CACHE_TTL_DETAIL_SECONDS", "120"))
redis_client = None


def init_redis_cache():
    """Initialize optional Redis cache for Home page APIs."""
    global redis_client
    redis_addr = (os.getenv("REDIS_ADDR") or "").strip()
    if not redis_addr:
        logger.info("Home Redis cache disabled: REDIS_ADDR is empty")
        return
    if redis is None:
        logger.warning("Home Redis cache disabled: python redis package not installed")
        return
    try:
        redis_db = int((os.getenv("REDIS_DB") or "0").strip() or "0")
    except Exception:
        redis_db = 0
    try:
        redis_client = redis.Redis(
            host=redis_addr.split(":")[0],
            port=int(redis_addr.split(":")[1]) if ":" in redis_addr else 6379,
            password=(os.getenv("REDIS_PASSWORD") or "").strip() or None,
            db=redis_db,
            socket_connect_timeout=1.2,
            socket_timeout=1.2,
            decode_responses=True,
        )
        redis_client.ping()
        logger.info("Home Redis cache enabled at %s (db=%s)", redis_addr, redis_db)
    except Exception as e:
        redis_client = None
        logger.warning("Home Redis cache disabled: ping failed: %s", str(e))


def _home_cache_enabled():
    return redis_client is not None


def _is_home_no_cache_request():
    no_cache = (request.args.get("nocache", "") or "").strip().lower()
    if no_cache in ("1", "true", "yes", "on"):
        return True
    force_refresh = (request.headers.get("X-Home-Force-Refresh", "") or "").strip()
    if force_refresh == "1":
        return True
    dashboard_force_refresh = (request.headers.get("X-Dashboard-Force-Refresh", "") or "").strip()
    return dashboard_force_refresh == "1"


def _is_apps_finder_no_cache_request():
    no_cache = (request.args.get("nocache", "") or "").strip().lower()
    if no_cache in ("1", "true", "yes", "on"):
        return True
    apps_finder_force_refresh = (request.headers.get("X-AppsFinder-Force-Refresh", "") or "").strip()
    if apps_finder_force_refresh == "1":
        return True
    dashboard_force_refresh = (request.headers.get("X-Dashboard-Force-Refresh", "") or "").strip()
    return dashboard_force_refresh == "1"


def _home_scope_key(current_user):
    team_id = (request.headers.get("X-Selected-Team-Id", "") or "").strip()
    if team_id:
        return f"team:{team_id}"
    user_id = current_user if isinstance(current_user, str) else (
        current_user.get("id") or current_user.get("user_id") or "self"
    )
    return f"user:{user_id}"


def _home_cache_generation():
    client = redis_client
    if client is None:
        return "0"
    try:
        val = client.get(HOME_CACHE_GENERATION_KEY)
        return str(val or "0")
    except Exception:
        return "0"


def bump_home_cache_generation(reason=""):
    client = redis_client
    if client is None:
        return
    try:
        new_val = client.incr(HOME_CACHE_GENERATION_KEY)
        logger.info("Home cache generation bumped to %s (reason=%s)", new_val, reason or "unknown")
    except Exception as e:
        logger.warning("Home cache generation bump failed: %s", str(e))


def _apps_finder_cache_generation():
    client = redis_client
    if client is None:
        return "0"
    try:
        val = client.get(APPS_FINDER_CACHE_GENERATION_KEY)
        return str(val or "0")
    except Exception:
        return "0"


def bump_apps_finder_cache_generation(reason=""):
    client = redis_client
    if client is None:
        return
    try:
        new_val = client.incr(APPS_FINDER_CACHE_GENERATION_KEY)
        logger.info("AppsFinder cache generation bumped to %s (reason=%s)", new_val, reason or "unknown")
    except Exception as e:
        logger.warning("AppsFinder cache generation bump failed: %s", str(e))


def _canonical_request_query():
    items = []
    for key in sorted(request.args.keys()):
        if key.lower() == "nocache":
            continue
        values = request.args.getlist(key)
        if not values:
            items.append((key, ""))
            continue
        for value in sorted(values):
            items.append((key, value))
    return urllib.parse.urlencode(items, doseq=True)


def _home_cache_key(namespace, current_user=None):
    generation = _home_cache_generation()
    scope = _home_scope_key(current_user) if current_user is not None else "anon"
    query = _canonical_request_query()
    raw_key = f"{HOME_CACHE_PREFIX}:{namespace}:{generation}:{scope}:{request.path}?{query}"
    digest = hashlib.sha256(raw_key.encode("utf-8")).hexdigest()
    return f"{HOME_CACHE_PREFIX}:{namespace}:{digest}"


def _apps_finder_cache_key(namespace, current_user=None):
    generation = _apps_finder_cache_generation()
    scope = _home_scope_key(current_user) if current_user is not None else "anon"
    query = _canonical_request_query()
    raw_key = f"{APPS_FINDER_CACHE_PREFIX}:{namespace}:{generation}:{scope}:{request.path}?{query}"
    digest = hashlib.sha256(raw_key.encode("utf-8")).hexdigest()
    return f"{APPS_FINDER_CACHE_PREFIX}:{namespace}:{digest}"


def _home_cache_get_json(cache_key):
    client = redis_client
    if client is None:
        return None
    try:
        payload = client.get(cache_key)
        if not payload:
            return None
        return json.loads(payload)
    except Exception:
        return None


def _home_cache_set_json(cache_key, value, ttl_seconds):
    client = redis_client
    if client is None or ttl_seconds <= 0:
        return
    try:
        jitter = max(1, int(ttl_seconds * 0.15))
        expire = max(1, ttl_seconds + (hash(cache_key) % (2 * jitter + 1)) - jitter)
        client.setex(cache_key, expire, json.dumps(value, ensure_ascii=False, default=str))
    except Exception as e:
        logger.warning("Home cache set failed: %s", str(e))


init_redis_cache()

# Log all registered routes
def print_routes():
    logger.debug('=== 已注册的路由 ===')
    for rule in app.url_map.iter_rules():
        logger.debug('路由: %s, 方法: %s', rule.rule, rule.methods)

# Create optimized SSL context
def create_ssl_context():
    """Create optimized SSL context."""
    context = ssl.create_default_context()
    # Set supported TLS versions
    context.minimum_version = ssl.TLSVersion.TLSv1_2
    context.maximum_version = ssl.TLSVersion.TLSv1_3
    # Set cipher suites
    context.set_ciphers('ECDHE+AESGCM:ECDHE+CHACHA20:DHE+AESGCM:DHE+CHACHA20:!aNULL:!MD5:!DSS')
    return context

# Register auth blueprint
app.register_blueprint(auth_bp)
logger.info("已注册认证蓝图")

# Configure CORS — comma-separated origins in CORS_ORIGINS (production: set in backend.env)
_cors_origins = [
    o.strip()
    for o in os.getenv(
        'CORS_ORIGINS',
        'http://localhost:3000,http://127.0.0.1:3000',
    ).split(',')
    if o.strip()
]
CORS(app, resources={
    r"/*": {
        "origins": _cors_origins,
        "methods": ["GET", "POST", "DELETE", "OPTIONS", "PUT", "PATCH"],
        "allow_headers": [
            "Content-Type", 
            "Authorization", 
            "Accept", 
            "Origin", 
            "X-Requested-With",
            "X-Environment",
            "X-API-Key",
            "X-Selected-Team-Id",
            "X-Dashboard-Force-Refresh",
            "X-Home-Force-Refresh",
            "X-AppsFinder-Force-Refresh"
        ],
        "expose_headers": [
            "Content-Type", 
            "Content-Disposition",
            "X-Environment",
            "X-API-Key",
            "X-Home-Cache",
            "X-AppsFinder-Cache"
        ],
        "supports_credentials": True,
        "max_age": 3600
    }
})

# Global OPTIONS handler
@app.route('/', defaults={'path': ''}, methods=['OPTIONS'])
@app.route('/<path:path>', methods=['OPTIONS'])
def handle_options(path):
    logger.debug('=== 处理OPTIONS请求 ===')
    logger.debug('请求路径: %s', path)
    logger.debug('完整URL: %s', request.url)
    logger.debug('请求头: %s', dict(request.headers))
    
    response = jsonify({'status': 'ok'})
    # Set CORS headers by environment
    if os.getenv('IS_LOCAL', 'false').lower() == 'true':
        # Local env: allow two origins
        origin = request.headers.get('Origin', '')
        if origin in ['http://localhost:3000', 'http://127.0.0.1:3000']:
            response.headers.add('Access-Control-Allow-Origin', origin)
        else:
            response.headers.add('Access-Control-Allow-Origin', 'http://localhost:3000')
    else:
        response.headers.add('Access-Control-Allow-Origin', os.getenv('CORS_ORIGIN', 'http://localhost:3000'))
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,POST,PUT,DELETE,OPTIONS')
    response.headers.add('Access-Control-Allow-Credentials', 'true')
    
    logger.debug('响应头: %s', dict(response.headers))
    return response

# Add DB test route
@app.route('/api/test-db', methods=['GET'])
def test_db_connection():
    """Test database connection."""
    logger.debug('=== Test database connection. ===')
    try:
        with get_db_cursor() as cursor:
            # Log connection info
            logger.info(f"数据库连接信息：")
            logger.info(f"主机：{os.getenv('DB_HOST', 'localhost')}")
            logger.info(f"环境：{'本地' if os.getenv('IS_LOCAL', 'false').lower() == 'true' else '生产'}")
            logger.info(f"连接类型：{'Unix Socket' if os.getenv('DB_HOST', 'localhost') == 'localhost' else 'TCP/IP'}")
            
            # Test users table
            cursor.execute("SELECT COUNT(*) as count FROM users")
            user_count = cursor.fetchone()['count']
            while cursor.nextset():
                pass
                
            # Test account config table
            cursor.execute("SELECT COUNT(*) as count FROM account_configs")
            config_count = cursor.fetchone()['count']
            while cursor.nextset():
                pass
                
            logger.debug('数据库连接测试成功')
            return jsonify({
                'status': 'success',
                'message': '数据库连接正常',
                'data': {
                    'users_count': user_count,
                    'configs_count': config_count,
                    'connection_info': {
                        'host': os.getenv('DB_HOST', 'localhost'),
                        'environment': '本地' if os.getenv('IS_LOCAL', 'false').lower() == 'true' else '生产',
                        'connection_type': 'Unix Socket' if os.getenv('DB_HOST', 'localhost') == 'localhost' else 'TCP/IP'
                    }
                }
            })
    except Exception as e:
        logger.error(f"数据库连接测试失败: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'数据库连接失败: {str(e)}'
        }), 500

# Initialize database
# init_db()  # Removed; init_db(app) already called above

# Initialize SocketIO
socketio = SocketIO(app, cors_allowed_origins="*", ping_timeout=60, ping_interval=25, logger=True, engineio_logger=True)

# Report data stored in DB

# App init flag
_app_initialized = False
_db_initialized = False

def check_db_connection():
    """Check database connection status."""
    global _db_initialized
    if _db_initialized:
        return True
        
    try:
        with get_db_cursor() as cursor:
            cursor.execute("SELECT 1")
            logger.info("数据库连接测试成功")
            _db_initialized = True
            return True
    except Exception as e:
        logger.error(f"数据库连接测试失败: {str(e)}")
        return False

def init_database():
    """Initialize database."""
    global _db_initialized
    if _db_initialized:
        return True
        
    try:
        with get_db_cursor() as cursor:
            # Create database if missing
            cursor.execute("CREATE DATABASE IF NOT EXISTS appsflyer_rawdata")
            cursor.execute("USE appsflyer_rawdata")
            
            # Create query_logs table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS query_logs (
                    id VARCHAR(100) PRIMARY KEY,
                    query_result_id VARCHAR(100) NOT NULL,
                    user_id VARCHAR(36) NOT NULL,
                    account_type VARCHAR(10) NOT NULL,
                    account_id VARCHAR(100) NOT NULL,
                    app_id VARCHAR(100) NOT NULL,
                    app_name VARCHAR(255),
                    data_type VARCHAR(50) NOT NULL,
                    from_date DATE NOT NULL,
                    to_date DATE NOT NULL,
                    status VARCHAR(20) NOT NULL,
                    message TEXT,
                    api_response JSON,
                    error_details JSON,
                    row_count INT,
                    event_filter VARCHAR(255),
                    mode ENUM('normal','aggregate') DEFAULT 'normal',
                    download_url VARCHAR(500) DEFAULT NULL COMMENT '下载URL',
                    afid_deduplication_count INT DEFAULT 0 COMMENT 'AppsFlyer ID去重数量',
                    primary_attribution_count INT DEFAULT 0 COMMENT 'Primary Attribution数量',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_user_id (user_id),
                    INDEX idx_query_result_id (query_result_id),
                    INDEX idx_account (account_type, account_id),
                    INDEX idx_created_at (created_at)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
            """)
            # Backfill missing columns on legacy DB (same as _ensure_query_logs_columns)
            _ensure_query_logs_columns(cursor)

            # Check if accounts table exists
            cursor.execute("SHOW TABLES LIKE 'accounts'")
            accounts_exists = cursor.fetchone() is not None
            
            if not accounts_exists:
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS accounts (
                        id VARCHAR(100) PRIMARY KEY,
                        account_type VARCHAR(50) NOT NULL,
                        account_name VARCHAR(255) NOT NULL,
                        account_id VARCHAR(100) NOT NULL,
                        api_token VARCHAR(255),
                        app_id VARCHAR(100),
                        app_name VARCHAR(255),
                        create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
                        update_time DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        INDEX idx_account_type (account_type),
                        INDEX idx_account_name (account_name),
                        INDEX idx_account_id (account_id)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
                """)
            
            logger.info("数据库表初始化成功")
            _db_initialized = True
            return True
    except Exception as e:
        logger.error(f"数据库初始化失败: {str(e)}")
        return False

def _ensure_query_logs_columns(cursor):
    """Ensure query_logs has mode etc. (legacy DB may lack columns) to avoid 1054 on INSERT/UPDATE. Uses current DB."""
    try:
        cursor.execute("SHOW COLUMNS FROM query_logs")
        _ql_cols = {r['Field'] for r in (cursor.fetchall() or [])}
        _ql_adds = []
        if 'app_name' not in _ql_cols:
            _ql_adds.append("ADD COLUMN app_name VARCHAR(255) DEFAULT NULL")
        if 'event_filter' not in _ql_cols:
            _ql_adds.append("ADD COLUMN event_filter VARCHAR(255) DEFAULT NULL")
        if 'mode' not in _ql_cols:
            _ql_adds.append("ADD COLUMN mode ENUM('normal','aggregate') DEFAULT 'normal'")
        if 'download_url' not in _ql_cols:
            _ql_adds.append("ADD COLUMN download_url VARCHAR(500) DEFAULT NULL")
        if 'afid_deduplication_count' not in _ql_cols:
            _ql_adds.append("ADD COLUMN afid_deduplication_count INT DEFAULT 0")
        if 'primary_attribution_count' not in _ql_cols:
            _ql_adds.append("ADD COLUMN primary_attribution_count INT DEFAULT 0")
        for _sql in _ql_adds:
            cursor.execute(f"ALTER TABLE query_logs {_sql}")
            logger.info("query_logs 已补充列: %s", _sql[:60])
    except Exception as e:
        logger.warning("query_logs 列检查/迁移: %s", e)


def ensure_temp_dir():
    """Ensure temp directory exists."""
    temp_dir = 'temp'
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
        logger.info(f"创建临时目录: {temp_dir}")

import os

# Use relative paths to avoid absolute-path issues
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMP_DIR = os.path.join(BASE_DIR, 'temp')

def init_app():
    """Initialize application."""
    global _app_initialized
    if _app_initialized:
        return True
    
    # Ensure temp dir exists
    ensure_temp_dir()
    
    # Test MySQL connection
    if not check_db_connection():
        logger.error("MySQL数据库连接测试失败")
        return False
    
    # Initialize MySQL tables
    if not init_database():
        logger.error("MySQL数据库表初始化失败")
        return False
    
    # PostgreSQL init handled by Go backend (ai_chat_service.go)
    # Skip PostgreSQL init; Gochat is fully in Go backend
    
    logger.info("应用初始化成功")
    _app_initialized = True
    return True

@app.before_request
def before_request():
    """Initialize app before each request."""
    if not _app_initialized:
        if not init_app():
            return jsonify({
                'status': 'error',
                'message': '应用初始化失败'
            }), 500

@socketio.on('connect')
def handle_connect():
    """Handle WebSocket connect."""
    try:
        logger.info('客户端已连接')
        # Send connect-success response
        emit('connect_response', {
            'status': 'success',
            'data': 'Connected',
            'timestamp': datetime.now().isoformat()
        })
        

    except Exception as e:
        error_msg = f'WebSocket连接处理失败: {str(e)}'
        logger.error(error_msg)
        emit('connect_response', {
            'status': 'error',
            'error': error_msg,
            'timestamp': datetime.now().isoformat()
        })

@socketio.on('disconnect')
def handle_disconnect(reason=None):
    """Handle WebSocket disconnect."""
    logger.info(f'Client disconnected: {reason}')

@socketio.on_error()
def handle_socket_error(e):
    """Handle WebSocket error."""
    logger.error(f'WebSocket错误: {str(e)}')

# Reports read from DB; no file storage








def ensure_account_temp_dir(account_type, account_id):
    """Ensure account temp directory exists."""
    temp_dir = os.path.join('temp', f"{account_type}_{account_id}")
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
        logger.info(f"创建账户临时目录: {temp_dir}")
    return temp_dir

def cleanup_orphaned_files():
    """Remove orphan files with no DB reference."""
    try:
        logger.info("开始清理孤立的物理文件")
        
        # List all files under temp
        # Relative path to backend/temp
        temp_dir = 'temp'
        if not os.path.exists(temp_dir):
            logger.info(f"temp目录不存在: {temp_dir}，跳过清理")
            return {
                'valid_files_count': 0,
                'total_files': 0,
                'deleted_files': 0,
                'retained_files': 0
            }
        
        # Step 1: files referenced in DB
        valid_files = set()
        with get_db_cursor() as cursor:
            # Get download_url from query_logs
            cursor.execute("""
                SELECT download_url FROM query_logs 
                WHERE download_url IS NOT NULL AND download_url != ''
            """)
            query_results = cursor.fetchall()
            
            for row in query_results:
                download_url = row['download_url']
                if download_url:
                    # Extract filename from URL; handle /api/download/ prefix
                    if download_url.startswith('/api/download/'):
                        filename = download_url.replace('/api/download/', '')
                    else:
                        filename = download_url.split('/')[-1]
                    valid_files.add(filename)
                    logger.info(f"query_logs有效文件: {filename}")
            

        
        valid_files_count = len(valid_files)
        logger.info(f"数据库中有引用的文件数量: {valid_files_count}")
        
        # Step 2: scan temp dir; delete orphan files
        cleaned_count = 0
        total_files = 0
        
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if file.endswith('.csv'):
                    total_files += 1
                    file_path = os.path.join(root, file)
                    
                    # Check file against valid list
                    if file in valid_files:
                        logger.info(f"保留有效文件: {file}")
                    else:
                        # Delete file not in valid list
                        try:
                            os.remove(file_path)
                            logger.info(f"删除无效文件: {file}")
                            cleaned_count += 1
                        except Exception as e:
                            logger.error(f"删除文件失败 {file_path}: {str(e)}")
        
        retained_files = total_files - cleaned_count
        logger.info(f"文件清理完成，总文件数: {total_files}, 删除: {cleaned_count}, 保留: {retained_files}")
        
        return {
            'valid_files_count': valid_files_count,
            'total_files': total_files,
            'deleted_files': cleaned_count,
            'retained_files': retained_files
        }
        
    except Exception as e:
        logger.error(f"文件清理过程中发生错误: {str(e)}")
        raise e

def check_and_cleanup_file(query_log_id):
    """Check and clean files when Home has no references."""
    try:
        with get_db_cursor() as cursor:
            # Check if query_logs has rows
            cursor.execute("""
                SELECT COUNT(*) as count FROM query_logs 
                WHERE id = %s OR query_result_id = %s
            """, (query_log_id, query_log_id))
            query_count = cursor.fetchone()['count']
            
            # No query_logs rows; safe to delete file
            if query_count == 0:
                # Resolve file path
                cursor.execute("""
                    SELECT api_response FROM query_logs 
                    WHERE id = %s OR query_result_id = %s
                """, (query_log_id, query_log_id))
                result = cursor.fetchone()
                
                if result and result['api_response']:
                    api_response = json.loads(result['api_response'])
                    download_url = api_response.get('downloadUrl', '')
                    
                    if download_url:
                        filename = download_url.split('/')[-1]
                        file_path = os.path.join('temp', filename)
                        
                        if os.path.exists(file_path):
                            try:
                                os.remove(file_path)
                                logger.info(f"删除物理文件: {file_path}")
                                return True
                            except Exception as e:
                                logger.error(f"删除物理文件失败: {str(e)}")
                                return False
        
        return False
        
    except Exception as e:
        logger.error(f"检查文件清理时发生错误: {str(e)}")
        return False





            










def decode_token(encoded_token):
    try:
        # Validate token format
        if not encoded_token or not isinstance(encoded_token, str):
            logger.warning('无效的token格式')
            return encoded_token
            
        # Return as-is if JWT (starts with eyJ)
        if encoded_token.startswith('eyJ'):
            logger.info('使用JWT格式token')
            return encoded_token
            
        # Try base64 decode
        try:
            # Check valid base64
            if len(encoded_token) % 4 == 0:
                decoded_token = base64.b64decode(encoded_token).decode('utf-8')
                logger.info(f'Token解码成功，长度: {len(decoded_token)}')
                return decoded_token
            else:
                logger.info('Token不是有效的base64格式，使用原始token')
                return encoded_token
        except Exception as e:
            logger.info(f'Base64解码失败，使用原始token: {str(e)}')
            return encoded_token
    except Exception as e:
        logger.error(f'Token处理失败: {str(e)}')
        return encoded_token

def fetch_data(account_name, account_type, data_type, from_date, to_date, app_id, api_token, event_filter=None, media_source='', mode='normal'):
    try:
        logger.info(f"Starting data fetch: account_name={account_name}, data_type={data_type}, app_id={app_id}, mode={mode}")
        
        # Validate API token
        if not api_token or len(api_token.strip()) == 0:
            logger.error("API Token is empty")
            return {
                'status': 'error',
                'message': 'API Token Cannot Be Empty',
                'details': {
                    'error_type': 'authentication',
                    'error_code': 'empty_token'
                }
            }

        # Decode token
        decoded_token = decode_token(api_token)
        
        # Log token metadata (not full token)
        token_length = len(decoded_token)
        token_prefix = decoded_token[:10] if token_length > 10 else decoded_token
        logger.info(f'Token length: {token_length}, prefix: {token_prefix}...')

        # Set request headers
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {decoded_token}'
        }

        # Build API request params
        if mode == 'aggregate':
            # Aggregate mode: PRT accounts only; simplified params
            params = {
                'from': from_date,
                'to': to_date
            }
        else:
            # Normal mode
            if account_type == 'PID':
                params = {
                    'from': from_date,
                    'to': to_date,
                    'media_source': 'standard',
                    'category': 'standard',
                    'install_type': 'organic',
                    'revenue': 'true',
                    'limit': '200000'
                }
                # Add event name param for filtered event types
                if (data_type == 'In-App-Event-Postbacks' or data_type == 'Retargeting-In-App-Event-Postbacks') and event_filter and event_filter.strip():
                    params['event_name'] = event_filter.strip()
            elif account_type == 'PRT':
                # PRT: four data types omit limit param
                prt_special_types = [
                    'In-App-Event-Non-Organic',
                    'Install-Non-Organic',
                    'Retargeting-In-App-Event-Non-Organic',
                    'Retargeting-Install-Non-Organic'
                ]
                
                params = {
                    'from': from_date,
                    'to': to_date
                }
                
                # Add limit only for non-PRT-special types
                if data_type not in prt_special_types:
                    params['limit'] = '200000'
                    logger.info(f'PRT账户非特殊类型，添加limit参数: {data_type}')
                else:
                    logger.info(f'PRT账户特殊类型，不添加limit参数: {data_type}')
                
                # Add media_source only when set and not 'All Media Source'
                # 'All Media Source' means unset; do not pass to API
                if media_source and media_source.strip() and media_source.strip() != 'All Media Source':
                    params['media_source'] = media_source.strip()
                    logger.info(f'PRT账户添加media_source参数: {media_source.strip()}')
                else:
                    logger.info(f'PRT账户未添加media_source参数，当前值: {media_source} (视为未指定)')
                # Add event name param for filtered event types
                if (data_type == 'In-App-Event-Non-Organic' or data_type == 'Retargeting-In-App-Event-Non-Organic') and event_filter and event_filter.strip():
                    params['event_name'] = event_filter.strip()
            else:
                raise ValueError(f'不支持的账户类型: {account_type}')

        # Prefix numeric app IDs with 'id'
        formatted_app_id = f"id{app_id}" if app_id.isdigit() else app_id
        logger.info(f'Formatted APP ID: {formatted_app_id}')

        # Pick API URL by account and data type
        url = None  # Initialize url
        
        if mode == 'aggregate':
            # Aggregate mode: PRT accounts only
            if account_type != 'PRT':
                raise ValueError(f'Aggregate mode only supports PRT account type, current account type: {account_type}')
            
            # Handle data types with "-Aggregate" suffix
            if data_type in ['daily', 'Daily-Aggregate']:
                url = f"https://hq1.appsflyer.com/api/agg-data/export/app/{formatted_app_id}/daily_report/v5"
            elif data_type in ['partner_daily', 'Partner-Daily-Aggregate']:
                url = f"https://hq1.appsflyer.com/api/agg-data/export/app/{formatted_app_id}/partners_by_date_report/v5"
            elif data_type in ['geo_daily', 'GEO-Daily-Aggregate']:
                url = f"https://hq1.appsflyer.com/api/agg-data/export/app/{formatted_app_id}/geo_by_date_report/v5"
            else:
                raise ValueError(f'Aggregate mode does not support data type: {data_type}')
        else:
            # Normal mode
            if account_type == 'PID':
                if data_type == 'In-App-Event-Postbacks':
                    url = f"https://hq1.appsflyer.com/api/raw-data/export/app/{formatted_app_id}/in-app-events-postbacks/v5"
                elif data_type == 'Install-Postbacks':
                    url = f"https://hq1.appsflyer.com/api/raw-data/export/app/{formatted_app_id}/postbacks/v5"
                elif data_type == 'Retargeting-In-App-Event-Postbacks':
                    url = f"https://hq1.appsflyer.com/api/raw-data/export/app/{formatted_app_id}/retarget_in_app_events_postbacks/v5"
                elif data_type == 'Retargeting-Install-Postbacks':
                    url = f"https://hq1.appsflyer.com/api/raw-data/export/app/{formatted_app_id}/retarget_install_postbacks/v5"
                else:
                    raise ValueError(f'PID账户不支持的数据类型: {data_type}')
            elif account_type == 'PRT':
                if data_type == 'In-App-Event-Non-Organic':
                    url = f"https://hq1.appsflyer.com/api/raw-data/export/app/{formatted_app_id}/in_app_events_report/v5"
                elif data_type == 'Install-Non-Organic':
                    url = f"https://hq1.appsflyer.com/api/raw-data/export/app/{formatted_app_id}/installs_report/v5"
                elif data_type == 'Retargeting-In-App-Event-Non-Organic':
                    url = f"https://hq1.appsflyer.com/api/raw-data/export/app/{formatted_app_id}/in-app-events-retarget/v5"
                elif data_type == 'Retargeting-Install-Non-Organic':
                    url = f"https://hq1.appsflyer.com/api/raw-data/export/app/{formatted_app_id}/installs-retarget/v5"
                else:
                    raise ValueError(f'PRT账户不支持的数据类型: {data_type}')
            else:
                raise ValueError(f'不支持的账户类型: {account_type}')
        
        # Ensure url is set
        if url is None:
            raise ValueError(f'无法为账户类型 {account_type} 和数据类型 {data_type} 构建URL')
        
        # Ensure headers is defined
        if 'headers' not in locals():
            headers = {
                'accept': 'application/json',
                'Authorization': f'Bearer {decoded_token}'
            }

        logger.info(f'发送API请求: {url}')
        logger.info(f'请求参数: {json.dumps(params, ensure_ascii=False)}')
        logger.info(f'请求头: {json.dumps({k: v for k, v in headers.items() if k != "Authorization"}, ensure_ascii=False)}')
        
        # Send API request with SSL and retries
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                # Create session with retry policy
                session = requests.Session()
                
                # Retry policy for HTTP status errors only
                retry_strategy = Retry(
                    total=2,  # Retries per session
                    backoff_factor=0.5,  # Retry backoff multiplier
                    status_forcelist=[429, 500, 502, 503, 504],  # HTTP status codes to retry
                    allowed_methods=["HEAD", "GET", "OPTIONS"]  # HTTP methods allowed to retry
                )
                
                # Mount HTTPAdapter with retry policy
                adapter = HTTPAdapter(max_retries=retry_strategy)  # type: ignore
                session.mount("http://", adapter)
                session.mount("https://", adapter)
                
                # Configure SSL
                session.verify = True  # Enable SSL verification
                # Note: session.timeout is not a standard requests.Session attribute
                # Set timeout on each request
                
                # Use optimized SSL context
                try:
                    ssl_context = create_ssl_context()
                    # HTTPAdapter has no ssl_context; use defaults
                    # session.mount('https://', HTTPAdapter())
                except Exception as ssl_config_error:
                    logger.warning(f'SSL上下文配置失败，使用默认配置: {str(ssl_config_error)}')
                
                # Suppress SSL warnings
                urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
                
                logger.info(f'尝试第 {retry_count + 1} 次请求到: {url}')
                response = session.get(url, params=params, headers=headers)
                
                # Break retry loop on success
                break
                
            except requests.exceptions.SSLError as ssl_error:
                retry_count += 1
                logger.error(f'SSL连接错误 (尝试 {retry_count}/{max_retries}): {str(ssl_error)}')
                
                if retry_count < max_retries:
                    # Wait before retry
                    import time
                    wait_time = retry_count * 2  # Increase wait between retries
                    logger.info(f'等待 {wait_time} 秒后重试...')
                    time.sleep(wait_time)
                    continue
                else:
                    # Final attempt: disable SSL verify
                    logger.warning('所有SSL验证尝试失败，尝试禁用SSL验证作为最后手段')
                    try:
                        session = requests.Session()
                        session.verify = False
                        # Timeout on request, not session
                        response = session.get(url, params=params, headers=headers)
                        logger.warning('使用禁用SSL验证成功连接')
                        break
                    except Exception as final_error:
                        logger.error(f'禁用SSL验证后仍然失败: {str(final_error)}')
                        raise Exception(f'SSL连接失败，已尝试所有方法: {str(ssl_error)}')
                        
            except requests.exceptions.ConnectionError as conn_error:
                retry_count += 1
                logger.error(f'连接错误 (尝试 {retry_count}/{max_retries}): {str(conn_error)}')
                
                if retry_count < max_retries:
                    import time
                    wait_time = retry_count * 2
                    logger.info(f'等待 {wait_time} 秒后重试...')
                    time.sleep(wait_time)
                    continue
                else:
                    raise Exception(f'无法连接到AppsFlyer API: {str(conn_error)}')
                    
            except requests.exceptions.Timeout as timeout_error:
                retry_count += 1
                logger.error(f'请求超时 (尝试 {retry_count}/{max_retries}): {str(timeout_error)}')
                
                if retry_count < max_retries:
                    import time
                    wait_time = retry_count * 2
                    logger.info(f'等待 {wait_time} 秒后重试...')
                    time.sleep(wait_time)
                    continue
                else:
                    raise Exception(f'请求AppsFlyer API超时: {str(timeout_error)}')
                    
            except Exception as e:
                retry_count += 1
                logger.error(f'请求AppsFlyer API时发生未知错误 (尝试 {retry_count}/{max_retries}): {str(e)}')
                
                if retry_count < max_retries:
                    import time
                    wait_time = retry_count * 2
                    logger.info(f'等待 {wait_time} 秒后重试...')
                    time.sleep(wait_time)
                    continue
                else:
                    raise Exception(f'请求AppsFlyer API失败: {str(e)}')
        
        # All retries exhausted
        if retry_count >= max_retries:
            raise Exception(f'请求AppsFlyer API失败，已尝试 {max_retries} 次')
        logger.info(f'AppsFlyer API 响应状态码: {response.status_code}')
        logger.info(f'AppsFlyer API 响应体: {response.text}')
        
        # Add detailed debug info
        logger.info(f'请求的日期范围: {from_date} 到 {to_date}')
        logger.info(f'请求的应用ID: {app_id} (格式化后: {formatted_app_id})')
        logger.info(f'Request data type: {data_type}')
        logger.info(f'Request account type: {account_type}')
        logger.info(f'Request mode: {mode}')
        

        
        if response.status_code != 200:
            error_details = {
                'status': response.status_code,
                'statusText': response.reason,
                'data': response.text,
                'config': {
                    'url': url,
                    'params': params,
                    'headers': {k: v for k, v in headers.items() if k != 'Authorization'}
                }
            }
            
            # Handle 416 Range Not Satisfiable
            if response.status_code == 416:
                logger.error('AppsFlyer返回416错误：范围无法满足')
                try:
                    error_data = json.loads(response.text)
                    error_message = error_data.get('data', response.text)
                except json.JSONDecodeError:
                    error_message = response.text
                
                return {
                    'status': 'error',
                    'message': 'Data Range Exceeded',
                    'details': {
                        'error_type': 'range_error',
                        'error_code': '416',
                        'error_message': error_message,
                        'config': error_details['config']
                    }
                }
            
            # Handle 404 (no authorization relationship)
            if response.status_code == 404:
                logger.error('无授权关系')
                return {
                    'status': 'error',
                    'message': 'No Authorization',
                    'details': {
                        'error_type': 'authorization',
                        'error_code': '404',
                        'error_message': 'No Authorization',
                        'config': error_details['config']
                    }
                }
            
            # Handle 400
            if response.status_code == 400:
                try:
                    # Try parse response text
                    response_text = response.text
                    logger.error(f'API返回400错误，原始响应: {response_text}')
                    
                    # Try parse JSON response
                    try:
                        error_data = json.loads(response_text)
                        # Prefer details.error_message; else data
                        error_message = error_data.get('details', {}).get('error_message', '') or error_data.get('data', '')
                    except json.JSONDecodeError:
                        error_message = response_text
                    
                    logger.error(f'解析后的错误消息: {error_message}')
                    
                    # Check subscription/raw-data permission error
                    if "your current subscription package doesn't include raw data reports" in error_message.lower():
                        error_response = {
                            'status': 'error',
                            'message': 'No Permission',
                            'details': {
                                'error_type': 'permission_error',
                                'error_code': '400',
                                'error_message': error_message,
                                'config': error_details['config'],
                                'log': {
                                    'account_type': account_type,
                                    'account_id': account_name,
                                    'app_id': app_id,
                                    'data_type': data_type,
                                    'from_date': from_date,
                                    'to_date': to_date,
                                    'status': 'failed',
                                    'message': 'No Permission'
                                }
                            }
                        }
                        logger.error(f"订阅包权限错误: {json.dumps(error_response, ensure_ascii=False)}")
                        return error_response
                    
                    # Check request limit error
                    if 'maximum number' in error_message.lower():
                        error_response = {
                            'status': 'error',
                            'message': 'Daily Request Limit Reached',
                            'details': {
                                'error_type': 'api_error',
                                'error_code': '400',
                                'error_message': error_message,
                                'config': error_details['config'],
                                'log': {
                                    'account_type': account_type,
                                    'account_id': account_name,
                                    'app_id': app_id,
                                    'data_type': data_type,
                                    'from_date': from_date,
                                    'to_date': to_date,
                                    'status': 'failed',
                                    'message': 'Daily Request Limit Reached'
                                }
                            }
                        }
                        logger.error(f"请求上限错误: {json.dumps(error_response, ensure_ascii=False)}")
                        return error_response
                    
                    # Extract day limit from error message
                    days_match = re.search(r'limited to (\d+) days', error_message)
                    if days_match:
                        days_limit = days_match.group(1)
                        error_response = {
                            'status': 'error',
                            'message': 'Time Range Limit',
                            'details': {
                                'error_type': 'api_error',
                                'error_code': '400',
                                'error_message': f'Data Limited to {days_limit} Days',
                                'config': error_details['config'],
                                'log': {
                                    'account_type': account_type,
                                    'account_id': account_name,
                                    'app_id': app_id,
                                    'data_type': data_type,
                                    'from_date': from_date,
                                    'to_date': to_date,
                                    'status': 'failed',
                                    'message': f'Data Limited to {days_limit} Days'
                                }
                            }
                        }
                        logger.error(f"时间范围限制错误: {json.dumps(error_response, ensure_ascii=False)}")
                        return error_response
                    else:
                        error_response = {
                            'status': 'error',
                            'message': 'Time Range Limit',
                            'details': {
                                'error_type': 'api_error',
                                'error_code': '400',
                                'error_message': error_message,
                                'config': error_details['config'],
                                'log': {
                                    'account_type': account_type,
                                    'account_id': account_name,
                                    'app_id': app_id,
                                    'data_type': data_type,
                                    'from_date': from_date,
                                    'to_date': to_date,
                                    'status': 'failed',
                                    'message': error_message
                                }
                            }
                        }
                        logger.error(f"API错误: {json.dumps(error_response, ensure_ascii=False)}")
                        return error_response
                except Exception as e:
                    logger.error(f'处理400错误时发生异常: {str(e)}')
                    return {
                        'status': 'error',
                        'message': 'Time Range Limit',
                        'details': {
                            'error_type': 'api_error',
                            'error_code': '400',
                            'error_message': 'Failed to Parse API Error Response',
                            'config': error_details['config']
                        }
                    }
            
            # Handle other non-200 errors
            # Try parse error response
            try:
                error_data = json.loads(response.text)
                error_message = error_data.get('data', response.text)
            except json.JSONDecodeError:
                error_message = response.text
            
            # Check permission-related errors
            response_text = response.text.lower()
            if any(keyword in response_text for keyword in ['unauthorized', 'forbidden', 'access denied', 'permission']):
                logger.error(f'权限错误: {error_message}')
                return {
                    'status': 'error',
                    'message': 'No Access Permission',
                    'details': {
                        'error_type': 'permission_error',
                        'error_code': str(response.status_code),
                        'error_message': error_message,
                        'config': error_details['config']
                    }
                }
            
            # Handle auth errors
            if response.status_code == 401:
                try:
                    error_data = json.loads(response.text)
                    error_message = error_data.get('error', 'Unknown Authentication Error')
                    logger.error(f'认证错误: {error_message}')
                    logger.error(f'Token信息: 长度={token_length}, 前缀={token_prefix}...')
                    return {
                        'status': 'error',
                        'message': f'Authentication Error: {error_message}',
                        'details': {
                            'error_type': 'authentication',
                            'error_code': 'invalid_token',
                            'error_message': error_message,
                            'token_info': {
                                'length': token_length,
                                'prefix': token_prefix
                            },
                            'config': error_details['config']
                        }
                    }
                except json.JSONDecodeError:
                    logger.error('无法解析错误响应')
                    return {
                        'status': 'error',
                        'message': 'Authentication Error',
                        'details': {
                            'error_type': 'authentication',
                            'error_code': 'invalid_token',
                            'error_message': 'Authentication Error',
                            'config': error_details['config']
                        }
                    }
            
            # Handle other errors
            logger.error(f'API请求失败: {response.status_code}')
            return {
                'status': 'error',
                'message': f'API Request Failed: {response.status_code}',
                'details': error_details
            }

        # Save response to temp file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Build new file naming scheme
        # 1. Prefix: RawData
        # 2. Add account_name
        # 3. Add filter tag for event types
        base_name = f"RawData({account_name})"
        
        # Check if event-type data
        is_event_type = data_type in ['In-App-Event-Postbacks', 'In-App-Event-Non-Organic', 'Retargeting-In-App-Event-Postbacks', 'Retargeting-In-App-Event-Non-Organic']
        
        # Add filter tag for event types
        if is_event_type:
            if event_filter and event_filter.strip():
                data_type_with_filter = f"{data_type}(Filter-On)"
            else:
                data_type_with_filter = f"{data_type}(Filter-Off)"
        else:
            data_type_with_filter = data_type
        
        # Build full filename
        if from_date != to_date:
            filename = f"{base_name}_{app_id}_{data_type_with_filter}_{from_date}——{to_date}_{timestamp}.csv"
        else:
            filename = f"{base_name}_{app_id}_{data_type_with_filter}_{from_date}_{timestamp}.csv"
        
        filepath = os.path.join('temp', filename)
        
        # Ensure temp dir exists
        os.makedirs('temp', exist_ok=True)
        
        # Process Postback URL fields (iOS only)
        processed_csv_content = response.text
        try:
            # Check iOS platform rows
            if 'platform' in response.text.lower() and 'ios' in response.text.lower():
                logger.info("检测到iOS平台数据，开始处理Postback URL字段")
                processed_csv_content = process_postback_urls_in_csv(response.text)
                logger.info("Postback URL字段处理完成")
        except Exception as e:
            logger.warning(f"处理Postback URL字段失败，使用原始数据: {str(e)}")
            processed_csv_content = response.text
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(processed_csv_content)
        
        logger.info(f'数据已保存到文件: {filepath}')

        # Count Is Primary Attribution for event types only
        primary_attribution_count = 0
        event_types = [
            'In-App-Event-Postbacks', 'Retargeting-In-App-Event-Postbacks',
            'In-App-Event-Non-Organic', 'Retargeting-In-App-Event-Non-Organic'
        ]
        
        if data_type in event_types:
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    # Read CSV header row
                    header_line = f.readline().strip()
                    csv_headers = header_line.split(',')
                    # Find Is Primary Attribution column (loose match)
                    primary_attribution_index = None
                    for i, header in enumerate(csv_headers):
                        normalized = header.replace(' ', '').replace('_', '').lower()
                        if normalized == 'isprimaryattribution':
                            primary_attribution_index = i
                            break
                    if primary_attribution_index is not None:
                        # Read rows and aggregate
                        for line in f:
                            values = [v.strip() for v in line.strip().split(',')]
                            if len(values) < len(csv_headers):
                                values += [''] * (len(csv_headers) - len(values))
                            if len(values) > primary_attribution_index:
                                logger.info(f"primary attribution values: {values}")
                                if values[primary_attribution_index].lower() == 'true':
                                    primary_attribution_count += 1
            except Exception as e:
                logger.error(f"统计Primary Attribution数量失败: {str(e)}")
        
        # Dedupe AppsFlyer ID count via csv.reader
        afid_deduplication_count = 0
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                csv_headers = next(reader)
                afid_index = None
                for i, header in enumerate(csv_headers):
                    if header.lower().replace('_', '').replace(' ', '') == 'appsflyerid':
                        afid_index = i
                        break
                if afid_index is not None:
                    afids = set()
                    for row in reader:
                        if len(row) > afid_index:
                            afid = row[afid_index].strip().replace('\r','').replace('\n','')
                            if afid:
                                afids.add(afid)
                    logger.info(f"去重AppsFlyer ID统计: {afids}")
                    afid_deduplication_count = len(afids)
        except Exception as e:
            logger.error(f"统计AppsFlyer ID去重数量失败: {str(e)}")

        # Get app_name from DB
        app_name = None
        try:
            with get_db_cursor() as cursor:
                cursor.execute("""
                    SELECT app_name FROM accounts 
                    WHERE account_type = %s AND account_id = %s AND app_id = %s
                    LIMIT 1
                """, (account_type, account_name, app_id))
                result = cursor.fetchone()
                if result:
                    app_name = result['app_name']
        except Exception as e:
            logger.error(f"从数据库获取app_name失败: {str(e)}")

        # If no app_name in DB, read from file
        if not app_name:
            try:
                # Read CSV header row
                with open(filepath, 'r', encoding='utf-8') as f:
                    header_line = f.readline().strip()
                    # Find app_name column
                    csv_headers = [h.strip() for h in header_line.split(',')]
                    app_name_index = None
                    for i, header in enumerate(csv_headers):
                        if header.lower() in ['app_name', 'appname', 'app name']:
                            app_name_index = i
                            break
                    if app_name_index is not None:
                        # Read first data row
                        data_line = f.readline().strip()
                        if data_line:
                            values = [v.strip() for v in data_line.split(',')]
                            # Pad row to header length
                            if len(values) < len(csv_headers):
                                values += [''] * (len(csv_headers) - len(values))
                            candidate = values[app_name_index]
                            logger.info(f"读取到 app_name 字段内容: {candidate}")
                            # Set app_name only when non-empty
                            if candidate:
                                app_name = candidate
                                # Persist app_name to DB
                                try:
                                    with get_db_cursor() as cursor:
                                        cursor.execute("""
                                            INSERT INTO accounts (
                                                id, account_type, account_name, account_id, 
                                                app_id, app_name, create_time, update_time
                                            ) VALUES (
                                                %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP
                                            ) AS new
                                            ON DUPLICATE KEY UPDATE 
                                                app_name = new.app_name,
                                                update_time = CURRENT_TIMESTAMP
                                        """, (
                                            str(uuid.uuid4()),
                                            account_type,
                                            account_name,
                                            account_name,
                                            app_id,
                                            app_name
                                        ))
                                        logger.info(f"已将app_name保存到数据库: {app_name}")
                                except Exception as e:
                                    logger.error(f"保存app_name到数据库失败: {str(e)}")
            except Exception as e:
                logger.error(f"从文件中读取app_name失败: {str(e)}")
        # Return None if app_name still empty
        if not app_name:
            # Try app_name from apps_finder
            try:
                with get_db_cursor() as cursor:
                    cursor.execute("SELECT app_name FROM apps_finder WHERE app_id = %s", (app_id,))
                    result = cursor.fetchone()
                    if result and result['app_name']:
                        app_name = result['app_name']
                        logger.info(f"从apps_finder获取到App Name: {app_name}")
            except Exception as e:
                logger.warning(f"从apps_finder查询App Name失败: {str(e)}")
        
        # Set app_name to None if still missing
        if not app_name:
            app_name = None

        # Build response payload
        response_data = {
            'status': 'success',
            'message': 'Data Fetched Successfully',
            'downloadUrl': f'/api/download/{filename}',
            'details': {
                'appId': app_id,
                'appName': app_name,
                'dataType': data_type,
                'dateRange': f'{from_date} To {to_date}',
                'rowCount': len(response.text.splitlines()) - 1,
                'afidDeduplicationCount': afid_deduplication_count,
                'config': {
                    'url': url,
                    'params': params,
                    'headers': {k: v for k, v in headers.items() if k != 'Authorization'}
                }
            }
        }
        
        # Add primaryAttributionCount for event types only
        if data_type in event_types:
            response_data['details']['primaryAttributionCount'] = primary_attribution_count

        return response_data
        
    except Exception as e:
        logger.error(f'获取数据失败: {str(e)}', exc_info=True)
        
        # Check network connection error
        error_str = str(e)
        if ('HTTPSConnectionPool' in error_str or 
            'SSLError' in error_str or 
            'SSL: UNEXPECTED_EOF_WHILE_READING' in error_str or
            'Max retries exceeded' in error_str or
            'ConnectionError' in error_str or
            'Timeout' in error_str):
            return {
                'status': 'error',
                'message': 'Network Connection Failed',
                'details': {
                    'error_type': 'network_error',
                    'error_code': 'NET_001',
                    'error_message': str(e),
                    'user_message': 'Network Connection Failed'
                }
            }
        else:
            return {
                'status': 'error',
                'message': f'Failed to Fetch Data: {str(e)}',
                'details': {
                    'error_type': 'system_error',
                    'error_code': 'SYS_001',
                    'error_message': str(e)
                }
            }

@app.route('/api/query-data', methods=['POST'])
@token_required
def query_data(current_user):
    """Execute data query."""
    logger.debug('=== Execute data query. ===')
    logger.debug('当前用户: %s', current_user)
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'status': 'error',
                'message': 'Invalid JSON Data',
                'queryId': None
            }), 400
            
        logger.debug('请求数据: %s', data)
        
        # Validate required fields
        required_fields = ['accountName', 'accountType', 'dataType', 'fromDate', 'toDate', 'appId', 'apiToken']
        for field in required_fields:
            if field not in data:
                return jsonify({
                    'status': 'error',
                    'message': f'Missing Required Field: {field}',
                    'queryId': None
                }), 400

        # Get event filter param
        event_filter = data.get('eventFilter')
        logger.debug('从请求数据中获取到事件过滤值: %s', event_filter)
        
        # Get media source as mediaSource (camelCase)
        media_source = data.get('mediaSource', '')
        # Treat empty or 'All Media Source' as unset; omit param
        if media_source == 'All Media Source' or not media_source or not media_source.strip():
            media_source = ''
        logger.debug('从请求数据中获取到mediaSource值: %s', media_source)
        
        # Get mode (Normal vs Aggregate)
        mode = data.get('mode', 'normal')
        logger.debug('Mode value retrieved from request data: %s', mode)

        # Ensure query_logs has mode etc. (legacy DB missing columns causes INSERT 500)
        with get_db_cursor() as cursor:
            _ensure_query_logs_columns(cursor)
        
        # Convert date format
        try:
            from_date = datetime.strptime(data['fromDate'], '%Y-%m-%d').date()
            to_date = datetime.strptime(data['toDate'], '%Y-%m-%d').date()
        except ValueError as e:
            return jsonify({
                'status': 'error',
                'message': f'Invalid Date Format: {str(e)}',
                'queryId': None
            }), 400

        # Server-side dedup: return existing row for same query; skip re-query
        with get_db_cursor() as cursor:
            cursor.execute("""
                SELECT id, status, message, api_response
                FROM query_logs
                WHERE user_id = %s
                  AND account_type = %s
                  AND account_id = %s
                  AND app_id = %s
                  AND data_type = %s
                  AND from_date = %s
                  AND to_date = %s
                  AND COALESCE(event_filter, '') = %s
                  AND COALESCE(mode, 'normal') = %s
                ORDER BY created_at DESC
                LIMIT 1
            """, (
                current_user,
                data['accountType'],
                data['accountName'],
                data['appId'],
                data['dataType'],
                from_date,
                to_date,
                event_filter or '',
                mode
            ))
            existing_record = cursor.fetchone()

        if existing_record:
            existing_api_response = existing_record.get('api_response')
            if isinstance(existing_api_response, str):
                try:
                    existing_api_response = json.loads(existing_api_response) if existing_api_response.strip() else {}
                except Exception:
                    existing_api_response = {}
            elif not isinstance(existing_api_response, dict):
                existing_api_response = {}

            return jsonify({
                'status': 'duplicate',
                'message': existing_record.get('message') or 'Duplicate query skipped',
                'queryId': existing_record['id'],
                'duplicate': True,
                'existingStatus': existing_record.get('status') or 'processing',
                'downloadUrl': existing_api_response.get('downloadUrl', ''),
                'details': existing_api_response.get('details', {}),
                'apiResponse': existing_api_response
            })
        
        # Generate query ID (short format + random suffix)
        timestamp = int(time.time())  # Strip milliseconds
        import random
        random_suffix = random.randint(1000, 9999)  # 4-digit random suffix
        query_id = f"{timestamp}_{random_suffix}_{data['appId']}"
        logger.debug('生成查询ID: %s', query_id)
        
        # Trim old rows over limit
        with get_db_cursor() as cursor:
            cursor.execute("""
                SELECT COUNT(*) as count FROM query_logs WHERE user_id = %s
            """, (current_user,))
            count = cursor.fetchone()['count']
            
            if count >= 30:
                # Delete oldest row
                cursor.execute("""
                    DELETE FROM query_logs 
                    WHERE user_id = %s 
                    AND id IN (
                        SELECT id FROM (
                            SELECT id FROM query_logs 
                            WHERE user_id = %s 
                            ORDER BY created_at ASC 
                            LIMIT 1
                        ) as temp
                    )
                """, (current_user, current_user))
                logger.debug('Deleted oldest query record')
        
        # Create query log
        with get_db_cursor() as cursor:
            # Check duplicate query_id
            cursor.execute("SELECT id FROM query_logs WHERE id = %s", (query_id,))
            if cursor.fetchone():
                # Regenerate query_id if duplicate
                import random
                random_suffix = random.randint(1000, 9999)
                query_id = f"{timestamp}_{random_suffix}_{data['appId']}"
                logger.debug('Regenerated query ID: %s', query_id)
            
            cursor.execute("""
                INSERT INTO query_logs (
                    id, query_result_id, user_id, account_type, account_id, 
                    app_id, app_name, data_type, from_date, to_date, status, message,
                    api_response, error_details, row_count, created_at,
                    event_filter, mode, afid_deduplication_count, primary_attribution_count
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, %s, %s, %s, %s
                )
            """, (
                query_id,
                query_id,  # query_result_id equals id
                current_user,
                data['accountType'],
                data['accountName'],
                data['appId'],
                None,  # app_name NULL initially; updated later
                data['dataType'],
                from_date,
                to_date,
                'pending',
                'Query started',
                json.dumps({}),  # Empty api_response initially
                json.dumps({}),  # Empty error_details initially
                0,  # Initial row_count
                event_filter,  # event_filter param
                mode,  # mode param
                0,  # Initial afid_deduplication_count
                0  # Initial primary_attribution_count
            ))
            logger.debug('Query log created successfully, event_filter: %s, mode: %s', event_filter, mode)
            bump_home_cache_generation("query_data:create_pending")
        
        # Run query
        try:
            result = fetch_data(
                data['accountName'],
                data['accountType'],
                data['dataType'],
                data['fromDate'],
                data['toDate'],
                data['appId'],
                data['apiToken'],
                event_filter,  # Event filter
                media_source,  # Use mediaSource variable
                mode  # Mode param
            )
            
            # Build full API response payload
            api_response = {
                'status': result.get('status'),
                'message': result.get('message'),
                'downloadUrl': result.get('downloadUrl'),
                'details': result.get('details', {}),
                'config': result.get('config', {}),
                'rowCount': result.get('row_count', 0)
            }
            
            # Resolve app name
            app_name = result.get('details', {}).get('appName')  # Fallback: appName from API response
            
            # Preferred: apps_finder lookup
            try:
                with get_db_cursor() as cursor:
                    cursor.execute("SELECT app_name FROM apps_finder WHERE app_id = %s", (data['appId'],))
                    result_app = cursor.fetchone()
                    if result_app and result_app['app_name']:
                        app_name = result_app['app_name']
                        logger.debug(f"从apps_finder获取到App Name: {app_name}")
            except Exception as e:
                logger.warning(f"从apps_finder查询App Name失败: {str(e)}")
            
            # Update query log
            with get_db_cursor() as cursor:
                # Build download URL
                download_url = None
                if result.get('status') == 'success' and result.get('downloadUrl'):
                    download_url = result.get('downloadUrl')
                
                cursor.execute("""
                    UPDATE query_logs 
                    SET status = %s,
                        message = %s,
                        api_response = %s,
                        error_details = %s,
                        row_count = %s,
                        app_name = %s,
                        download_url = %s,
                        event_filter = %s,
                        mode = %s,
                        afid_deduplication_count = %s,
                        primary_attribution_count = %s
                    WHERE id = %s
                """, (
                    'success' if result.get('status') == 'success' else 'failed',
                    result.get('message', 'Query Completed'),
                    json.dumps(api_response),
                    json.dumps(result.get('error_details', {})),
                    result.get('row_count', 0),
                    app_name,  # Use resolved app name
                    download_url,
                    event_filter,  # Use event_filter instead of data.get('eventFilter')
                    mode,  # Add mode field
                    result.get('details', {}).get('afidDeduplicationCount', 0),  # afid_deduplication_count
                    result.get('details', {}).get('primaryAttributionCount', 0),  # primary_attribution_count
                    query_id
                ))
                logger.debug('更新查询日志成功, event_filter: %s, app_name: %s', event_filter, app_name)
                bump_home_cache_generation("query_data:update_result")
            
            # Ensure queryId in response
            result['queryId'] = query_id
            return jsonify(result)
            
        except Exception as e:
            error_details = {
                'error': str(e),
                'timestamp': datetime.now().isoformat(),
                'query_id': query_id
            }
            
            # Mark query log failed
            with get_db_cursor() as cursor:
                cursor.execute("""
                    UPDATE query_logs 
                    SET status = 'failed',
                        message = %s,
                        error_details = %s
                    WHERE id = %s
                """, (
                    str(e),
                    json.dumps(error_details),
                    query_id
                ))
                logger.debug('更新查询日志为失败状态')
                bump_home_cache_generation("query_data:update_failed")
            
            # Return error with queryId
            return jsonify({
                'status': 'error',
                'message': f'Failed to Query Data: {str(e)}',
                'queryId': query_id,
                'error_details': error_details
            }), 500
            
    except Exception as e:
        logger.error(f"查询数据失败: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'查询数据失败: {str(e)}',
            'queryId': None
        }), 500

@app.route('/api/download/<filename>')
def download_file(filename):
    try:
        logger.info(f"开始处理文件下载请求: {filename}")
        
        # Resolve path; prefer relative
        possible_paths = [
            os.path.join('temp', filename),  # Relative path first
            os.path.join(TEMP_DIR, filename),  # Absolute path fallback
            filename  # Bare filename last resort
        ]
        
        file_path = None
        logger.info(f"开始查找文件: {filename}")
        logger.info(f"当前工作目录: {os.getcwd()}")
        logger.info(f"BASE_DIR: {BASE_DIR}")
        logger.info(f"TEMP_DIR: {TEMP_DIR}")
        
        for path in possible_paths:
            logger.info(f"检查路径: {path} (存在: {os.path.exists(path)})")
            if os.path.exists(path):
                file_path = path
                logger.info(f"找到文件: {path}")
                break
        
        if not file_path:
            logger.error(f"文件不存在: {filename}")
            logger.error(f"尝试的路径: {possible_paths}")
            
            # List temp dir files for debug
            try:
                temp_dirs = ['temp', TEMP_DIR]
                for temp_dir in temp_dirs:
                    if os.path.exists(temp_dir):
                        logger.info(f"检查目录: {temp_dir}")
                        for root, dirs, files in os.walk(temp_dir):
                            for file in files:
                                if file.endswith('.csv'):
                                    logger.info(f"找到CSV文件: {os.path.join(root, file)}")
            except Exception as e:
                logger.error(f"列出文件时出错: {str(e)}")
            
            return jsonify({
                'status': 'error',
                'message': '文件不存在'
            }), 404
        
        # Check file readable
        if not os.access(file_path, os.R_OK):
            logger.error(f"文件不可读: {file_path}")
            return jsonify({
                'status': 'error',
                'message': 'File Not Readable'
            }), 403
        
        # Get file size
        file_size = os.path.getsize(file_path)
        logger.info(f"文件大小: {file_size} 字节")
        
        # Set response headers
        response = send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv',
            conditional=True
        )
        
        # Add required headers
        response.headers['Content-Length'] = file_size
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response.headers['Cache-Control'] = 'no-cache'
        response.headers['X-Content-Type-Options'] = 'nosniff'
        
        # Encode filename for Content-Disposition
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        response.headers['Content-Disposition'] = f'attachment; filename="{encoded_filename}"; filename*=UTF-8\'\'{encoded_filename}'
        
        logger.info(f"文件下载响应已准备: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"文件下载失败: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Failed to Download File: {str(e)}'
        }), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    try:
        # Check DB connection
        with get_db_cursor() as cursor:
            cursor.execute("SELECT 1")
            # Get pool status
            pool_status = {
                "pool_size": connection_pool.pool_size,
                "pool_name": connection_pool.pool_name
            }
            
        return jsonify({
            "status": "healthy",
            "database": {
                "connected": True,
                "pool_status": pool_status,
                "connection_info": {
                    "host": os.getenv('DB_HOST', 'localhost'),
                    "environment": "本地" if os.getenv('IS_LOCAL', 'false').lower() == 'true' else "生产",
                    "connection_type": "Unix Socket" if os.getenv('DB_HOST', 'localhost') == 'localhost' else "TCP/IP"
                }
            },
            "environment": os.getenv('FLASK_ENV', 'development'),
            "timestamp": datetime.now().isoformat()
        })
    except Exception as e:
        logger.error(f"健康检查失败: {str(e)}")
        return jsonify({
            "status": "unhealthy",
            "error": str(e),
            "environment": os.getenv('FLASK_ENV', 'development'),
            "timestamp": datetime.now().isoformat()
        }), 500

@app.route('/api/cleanup-orphaned-files', methods=['POST'])
def cleanup_orphaned_files_api():
    """API endpoint to manually clean orphan files."""
    try:
        # Get cleanup stats
        stats = cleanup_orphaned_files()
        return jsonify({
            'status': 'success',
            'message': 'Files cleanup completed successfully',
            'stats': stats
        })
    except Exception as e:
        logger.error(f"清理孤立文件失败: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Cleanup failed: {str(e)}'
        }), 500

@app.route('/api/ping', methods=['GET'])
def ping():
    """Check AppsFlyer API connectivity."""
    try:
        start_time = datetime.now()
        # Use simpler connectivity check
        try:
            # Probe simpler API endpoint
            response = requests.get(
                'https://hq1.appsflyer.com/api/v1/status',
                timeout=10,  # Increase timeout
                verify=True,  # Enable SSL verification
                headers={'Accept': 'application/json'},
                allow_redirects=True
            )
            
            end_time = datetime.now()
            ping_time = int((end_time - start_time).total_seconds() * 1000)
            
            # Treat non-404 as reachable
            if response.status_code != 404:
                logger.info(f'Ping检测成功: {ping_time}ms')
                return jsonify({
                    'success': True,
                    'pingTime': ping_time,
                    'status': 'good' if ping_time < 1000 else 'warning' if ping_time < 2000 else 'poor'
                })
            else:
                logger.error(f'Ping检测失败: HTTP 404')
                return jsonify({
                    'success': False,
                    'pingTime': 3000,
                    'status': 'poor',
                    'error': 'Network Error'
                })
                
        except requests.exceptions.RequestException as e:
            logger.error(f'Ping检测失败: {str(e)}')
            return jsonify({
                'success': False,
                'pingTime': 3000,
                'status': 'poor',
                'error': str(e)
            })
            
    except Exception as e:
        logger.error(f'Ping检测失败: {str(e)}')
        return jsonify({
            'success': False,
            'pingTime': 3000,
            'status': 'poor',
            'error': str(e)
        })






# Log request handling
@app.before_request
def log_request_info():
    logger.debug('=== 收到新请求 ===')
    logger.debug('请求路径: %s', request.path)
    logger.debug('请求方法: %s', request.method)
    logger.debug('请求头: %s', dict(request.headers))
    logger.debug('请求URL: %s', request.url)
    logger.debug('请求参数: %s', request.args)
    logger.debug('请求数据: %s', request.get_data())
    logger.debug('请求JSON: %s', request.get_json(silent=True))
    logger.debug('请求表单: %s', request.form)
    logger.debug('请求文件: %s', request.files)
    logger.debug('请求环境: %s', dict(request.environ))
    logger.debug('注册的路由: %s', [str(rule) for rule in app.url_map.iter_rules()])

# Log response handling
@app.after_request
def log_response_info(response):
    logger.debug('=== 发送响应 ===')
    logger.debug('响应状态: %s', response.status)
    logger.debug('响应头: %s', dict(response.headers))
    
    # Check file-download response
    if 'Content-Disposition' in response.headers and 'attachment' in response.headers['Content-Disposition']:
        logger.debug('文件下载响应，跳过数据记录')
    else:
        try:
            logger.debug('响应数据: %s', response.get_data())
        except Exception as e:
            logger.debug('无法获取响应数据: %s', str(e))
    
    return response

# Query result API routes
@app.route('/api/query-results', methods=['GET'])
@token_required
def get_query_results(current_user):
    """List query results."""
    logger.debug('=== List query results. ===')
    logger.debug('当前用户: %s', current_user)
    
    try:
        cache_key = None
        if _home_cache_enabled() and not _is_home_no_cache_request():
            cache_key = _home_cache_key("query-results", current_user=current_user)
            cached_payload = _home_cache_get_json(cache_key)
            if cached_payload is not None:
                response = jsonify(cached_payload)
                response.headers["X-Home-Cache"] = "HIT"
                return response

        # Get query params
        mode = request.args.get('mode', 'all')  # Default: all modes
        logger.debug('Request mode filter: %s', mode)

        with get_db_cursor() as cursor:
            # Team scope: no team = current user only; X-Selected-Team-Id = that team's members (excl. Super Admin unless member)
            effective_ids = list(_effective_user_ids_for_data_scope(current_user, cursor))
            in_placeholders = ','.join(['%s'] * len(effective_ids))
            in_params = tuple(effective_ids)

            # Filter by mode
            if mode == 'normal':
                cursor.execute(f"""
                    SELECT 
                        id,
                        query_result_id,
                        user_id,
                        account_type,
                        account_id,
                        app_id,
                        app_name,
                        data_type,
                        from_date,
                        to_date,
                        status,
                        message,
                        api_response,
                        error_details,
                        row_count,
                        event_filter,
                        mode,
                        afid_deduplication_count,
                        created_at
                    FROM query_logs 
                    WHERE user_id IN ({in_placeholders}) AND mode = 'normal'
                    ORDER BY created_at DESC
                    LIMIT 30
                """, in_params)
            elif mode == 'aggregate':
                cursor.execute(f"""
                    SELECT 
                        id,
                        query_result_id,
                        user_id,
                        account_type,
                        account_id,
                        app_id,
                        app_name,
                        data_type,
                        from_date,
                        to_date,
                        status,
                        message,
                        api_response,
                        error_details,
                        row_count,
                        event_filter,
                        mode,
                        afid_deduplication_count,
                        created_at
                    FROM query_logs 
                    WHERE user_id IN ({in_placeholders}) AND mode = 'aggregate'
                    ORDER BY created_at DESC
                    LIMIT 30
                """, in_params)
            else:
                cursor.execute(f"""
                    SELECT 
                        id,
                        query_result_id,
                        user_id,
                        account_type,
                        account_id,
                        app_id,
                        app_name,
                        data_type,
                        from_date,
                        to_date,
                        status,
                        message,
                        api_response,
                        error_details,
                        row_count,
                        event_filter,
                        mode,
                        afid_deduplication_count,
                        created_at
                    FROM query_logs 
                    WHERE user_id IN ({in_placeholders})
                    ORDER BY created_at DESC
                    LIMIT 30
                """, in_params)

            results = cursor.fetchall()
            logger.debug('查询到 %d 条记录', len(results))

            def _safe_json_field(val):
                if val is None:
                    return {}
                if isinstance(val, dict):
                    return val
                if isinstance(val, (bytes, bytearray)):
                    try:
                        val = val.decode('utf-8')
                    except Exception:
                        return {}
                if isinstance(val, str):
                    try:
                        return json.loads(val) if val.strip() else {}
                    except (json.JSONDecodeError, TypeError):
                        return {}
                return {}

            # Debug: log each record
            for i, result in enumerate(results):
                logger.debug('Record %d: id=%s, mode=%s, status=%s, app_id=%s, data_type=%s',
                           i+1, result['id'], result.get('mode'), result['status'],
                           result['app_id'], result['data_type'])

            # Format results
            formatted_results = []
            status_priority = {
                'success': 5,
                'processing': 4,
                'pending': 3,
                'failed': 2,
                'error': 1
            }
            # Server dedup by business key; keep higher-priority or newer row
            dedup_map = {}
            for result in results:
                # Parse API response (MySQL JSON may already be dict; skip json.loads)
                api_response = _safe_json_field(result.get('api_response'))
                error_details = _safe_json_field(result.get('error_details'))
                
                # Build full response payload
                formatted_result = {
                    'key': result['id'],
                    'queryResultId': result['query_result_id'],
                    'userId': result['user_id'],
                    'accountType': result['account_type'],
                    'accountId': result['account_id'],
                    'appId': result['app_id'],
                    'appName': result.get('app_name'),
                    'dataType': result['data_type'],
                    'dateRange': f"{result['from_date']} To {result['to_date']}",
                    'status': result['status'],
                    'message': result['message'],
                    'apiResponse': api_response,
                    'errorDetails': error_details,
                    'rowCount': result['row_count'],
                    'event_filter': result['event_filter'],
                    'mode': result.get('mode') or 'normal',
                    'afidDeduplicationCount': result.get('afid_deduplication_count', 0),
                    'createTime': (
                        result['created_at'].strftime('%Y-%m-%d %H:%M:%S')
                        if result.get('created_at') and hasattr(result['created_at'], 'strftime')
                        else (str(result['created_at']) if result.get('created_at') else '')
                    ),
                    'downloadUrl': api_response.get('downloadUrl', ''),
                    'details': api_response.get('details', {}),
                    'config': api_response.get('config', {})
                }

                dedup_key = "|".join([
                    str(result.get('user_id') or ''),
                    str(result.get('account_type') or ''),
                    str(result.get('account_id') or ''),
                    str(result.get('app_id') or ''),
                    str(result.get('data_type') or ''),
                    str(result.get('from_date') or ''),
                    str(result.get('to_date') or ''),
                    str(result.get('event_filter') or ''),
                    str(result.get('mode') or 'normal')
                ])

                existed = dedup_map.get(dedup_key)
                current_priority = status_priority.get(str(result.get('status') or '').lower(), 0)
                current_time = result.get('created_at')
                if not existed:
                    dedup_map[dedup_key] = {
                        'priority': current_priority,
                        'created_at': current_time,
                        'record': formatted_result
                    }
                else:
                    existed_priority = existed['priority']
                    existed_time = existed['created_at']
                    should_replace = False
                    if current_priority > existed_priority:
                        should_replace = True
                    elif current_priority == existed_priority:
                        if current_time and existed_time:
                            should_replace = current_time > existed_time
                        elif current_time and not existed_time:
                            should_replace = True
                    if should_replace:
                        dedup_map[dedup_key] = {
                            'priority': current_priority,
                            'created_at': current_time,
                            'record': formatted_result
                        }

            formatted_results = [item['record'] for item in dedup_map.values()]
            formatted_results.sort(
                key=lambda x: x.get('createTime') or '',
                reverse=True
            )
            
            payload = {
                'status': 'success',
                'data': formatted_results
            }
            if cache_key:
                _home_cache_set_json(cache_key, payload, HOME_CACHE_TTL_QUERY_RESULTS)
            response = jsonify(payload)
            response.headers["X-Home-Cache"] = "MISS" if cache_key else "BYPASS"
            return response
            
    except Exception as e:
        logger.error(f"获取查询结果失败: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to Get Query Results: {str(e)}'
        }), 500

@app.route('/api/query-results/<result_key>', methods=['PUT'])
@token_required
def update_query_result(current_user, result_key):
    """Update query result record."""
    try:
        data = request.get_json()
        
        # Validate account
        account_type = data.get('accountType')
        account_id = data.get('accountId')
        if not account_type or not account_id:
            return jsonify({"error": "Missing Account Information"}), 400
        
        # Update record
        with get_db_cursor() as cursor:
            cursor.execute("""
                UPDATE query_results 
                SET status = %(status)s,
                    message = %(message)s,
                    download_url = %(download_url)s,
                    api_response = %(api_response)s,
                    error_details = %(error_details)s,
                    app_name = %(app_name)s,
                    primary_attribution_count = %(primary_attribution_count)s
                WHERE key = %(key)s 
                AND account_type = %(account_type)s 
                AND account_id = %(account_id)s
            """, {
                'key': result_key,
                'status': data.get('status'),
                'message': data.get('message'),
                'download_url': data.get('downloadUrl'),
                'api_response': json.dumps(data.get('apiResponse', {})),
                'error_details': json.dumps(data.get('errorDetails', {})),
                'app_name': data.get('appName'),
                'primary_attribution_count': data.get('primaryAttributionCount', 0),
                'account_type': account_type,
                'account_id': account_id
            })
            
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"更新查询结果失败: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/api/query-results/<result_key>', methods=['DELETE'])
@token_required
def delete_query_result(current_user, result_key):
    """Delete query result record."""
    try:
        logger.info(f"开始删除查询结果: {result_key}")
        
        # Get account from request
        account_type = request.args.get('accountType')
        account_id = request.args.get('accountId')
        
        logger.info(f"账户信息: accountType={account_type}, accountId={account_id}")
        logger.info(f"所有请求参数: {dict(request.args)}")
        
        if not account_type or not account_id:
            return jsonify({"error": "Missing Account Information"}), 400
        
        # Fetch record incl. downloadUrl
        with get_db_cursor() as cursor:
            # Add SQL debug logging
            logger.info(f"SQL查询参数: result_key={result_key}, account_type={account_type}, account_id={account_id}")
            logger.info(f"SQL查询参数类型: result_key={type(result_key)}, account_type={type(account_type)}, account_id={type(account_id)}")
            
            # Try simple lookup first
            cursor.execute("""
                SELECT id, account_type, account_id, download_url, app_id, data_type, from_date, to_date 
                FROM query_logs 
                WHERE `id` = %s
            """, (result_key,))
            
            result = cursor.fetchone()
            if not result:
                logger.warning(f"记录不存在: {result_key}")
                return jsonify({"error": "记录不存在"}), 404
            
            # Verify account matches
            db_account_type = result.get('account_type')
            db_account_id = result.get('account_id')
            
            logger.info(f"数据库中的账户信息: account_type={db_account_type}, account_id={db_account_id}")
            logger.info(f"请求中的账户信息: account_type={account_type}, account_id={account_id}")
            
            if db_account_type != account_type or db_account_id != account_id:
                logger.warning(f"账户信息不匹配！数据库: {db_account_type}/{db_account_id}, 请求: {account_type}/{account_id}")
                return jsonify({"error": "账户信息不匹配"}), 403
            
            download_url = result.get('download_url')
            app_id = result.get('app_id')
            data_type = result.get('data_type')
            from_date = result.get('from_date')
            to_date = result.get('to_date')
            
            logger.info(f"找到记录: app_id={app_id}, data_type={data_type}, from_date={from_date}, to_date={to_date}")
            logger.info(f"download_url: {download_url}")
            
            # Delete file first; skip DB delete if file delete fails
            file_deleted = True  # Assume file delete OK when no file
            
            if download_url:
                try:
                    # Extract filename from downloadUrl
                    # download_url may be full path e.g. /api/download/filename.csv
                    if download_url.startswith('/api/download/'):
                        filename = download_url.replace('/api/download/', '')
                    elif '/' in download_url:
                        filename = download_url.split('/')[-1]
                    else:
                        filename = download_url
                    
                    logger.info(f"尝试删除文件: {filename}, 原始download_url: {download_url}")
                    
                    # Candidate file paths
                    possible_paths = [
                        os.path.join('temp', filename),  # Relative path first
                        os.path.join(os.getcwd(), 'temp', filename),  # Absolute path
                        filename  # Bare filename last resort
                    ]
                    
                    file_deleted = False
                    for path in possible_paths:
                        if os.path.exists(path) and os.path.isfile(path):
                            try:
                                os.remove(path)
                                logger.info(f"成功删除物理文件: {path}")
                                file_deleted = True
                                break
                            except OSError as e:
                                logger.warning(f"删除物理文件失败 {path}: {str(e)}")
                                continue
                    
                    if not file_deleted:
                        logger.warning(f"未找到或无法删除物理文件: {filename}")
                        # List temp dir files for debug
                        try:
                            temp_dirs = ['temp', os.path.join(os.getcwd(), 'temp')]
                            for temp_dir in temp_dirs:
                                if os.path.exists(temp_dir):
                                    logger.info(f"temp目录 {temp_dir} 下的文件:")
                                    for file in os.listdir(temp_dir):
                                        if os.path.isfile(os.path.join(temp_dir, file)):
                                            logger.info(f"  - {file}")
                        except Exception as list_error:
                            logger.error(f"列出temp目录文件失败: {str(list_error)}")
                        
                except Exception as file_error:
                    logger.error(f"删除物理文件时出错: {str(file_error)}")
                    file_deleted = False
            else:
                logger.info("记录没有download_url，跳过物理文件删除")
            
            # Delete DB row only if file gone or absent
            if file_deleted:
                cursor.execute("""
                    DELETE FROM query_logs 
                    WHERE `id` = %s
                """, (result_key,))
                
                deleted_rows = cursor.rowcount
                logger.info(f"数据库删除结果: {deleted_rows} 行被删除")
            else:
                logger.error(f"物理文件删除失败，不删除数据库记录: {result_key}")
                return jsonify({"error": "Failed to Delete Physical File, Cannot Delete Record"}), 500
            
        logger.info(f"删除查询结果完成: {result_key}")
        bump_home_cache_generation("query_result:delete")
        return jsonify({"success": True, "deleted_rows": deleted_rows})
    except Exception as e:
        logger.error(f"删除查询结果失败: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/api/delete-file/<filename>', methods=['DELETE'])
@token_required
def delete_file(current_user, filename):
    """Delete physical file under temp."""
    try:
        # Reject path traversal in filename
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({
                'status': 'error',
                'message': 'Invalid filename'
            }), 400
        
        # Candidate file paths
        possible_paths = [
            os.path.join('temp', filename),  # Relative path first
            os.path.join(os.getcwd(), 'temp', filename),  # Absolute path
            filename  # Bare filename last resort
        ]
        
        file_deleted = False
        for path in possible_paths:
            if os.path.exists(path) and os.path.isfile(path):
                try:
                    os.remove(path)
                    logger.info(f"成功删除文件: {path}")
                    file_deleted = True
                    break
                except OSError as e:
                    logger.warning(f"删除文件失败 {path}: {str(e)}")
                    continue
        
        if file_deleted:
            return jsonify({
                'status': 'success',
                'message': f'File {filename} deleted successfully'
            })
        else:
            return jsonify({
                'status': 'warning',
                'message': f'File {filename} not found or could not be deleted'
            }), 404
            
    except Exception as e:
        logger.error(f"删除文件失败: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/api/query-logs', methods=['POST'])
@token_required
def create_query_log(current_user):
    try:
        data = request.get_json()
        logger.debug(f"创建查询日志: {data}")
        
        # DB connection via with
        with get_db_cursor() as cursor:
            # Insert query log
            cursor.execute("""
                INSERT INTO query_logs (
                    id,
                    query_result_id,
                    user_id,
                    account_type,
                    account_id,
                    app_id,
                    app_name,
                    data_type,
                    from_date,
                    to_date,
                    status,
                    message,
                    api_response,
                    error_details,
                    row_count,
                    created_at,
                    event_filter
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, %s
                )
            """, (
                data.get('queryResultId'),
                data.get('queryResultId'),
                current_user,
                data.get('accountType'),
                data.get('accountId'),
                data.get('appId'),
                data.get('appName'),
                data.get('dataType'),
                data.get('fromDate'),
                data.get('toDate'),
                data.get('status', 'pending'),
                data.get('message', ''),
                json.dumps(data.get('apiResponse', {})),
                json.dumps(data.get('errorDetails', {})),
                data.get('rowCount', 0),
                data.get('eventFilter', None)
            ))
            
            # Commit transaction
            cursor.connection.commit()
            
            logger.debug(f"查询日志创建成功: {data.get('queryResultId')}")
            bump_home_cache_generation("query_log:create")
            return jsonify({'message': 'Query log created successfully'})
            
    except Exception as e:
        logger.error(f"创建查询日志失败: {str(e)}")
        return jsonify({'message': str(e)}), 500

@app.route('/api/query-logs/<query_result_id>', methods=['PUT'])
@token_required
def update_query_log(current_user, query_result_id):
    try:
        data = request.get_json()
        logger.debug(f"更新查询日志: {query_result_id}, 数据: {data}")
        
        # Convert date format
        try:
            from_date = None
            to_date = None
            if data.get('fromDate'):
                from_date = datetime.strptime(data.get('fromDate'), '%Y-%m-%d').date()
            if data.get('toDate'):
                to_date = datetime.strptime(data.get('toDate'), '%Y-%m-%d').date()
        except (ValueError, TypeError) as e:
            logger.error(f"日期格式转换失败: {str(e)}")
            return jsonify({
                'status': 'error',
                'message': f'Invalid Date Format: {str(e)}'
            }), 400
        
        # DB connection via with
        with get_db_cursor() as cursor:
            # Check record exists
            cursor.execute("""
                SELECT id FROM query_logs 
                WHERE id = %s OR query_result_id = %s
            """, (query_result_id, query_result_id))
            existing_log = cursor.fetchone()
            
            if not existing_log:
                logger.warning(f"未找到要更新的查询日志: {query_result_id}")
                # Create row if missing
                if not from_date or not to_date:
                    return jsonify({
                        'status': 'error',
                        'message': 'fromDate and toDate Cannot Be Empty When Creating New Record'
                    }), 400
                    
                cursor.execute("""
                    INSERT INTO query_logs (
                        id,
                        query_result_id,
                        user_id,
                        account_type,
                        account_id,
                        app_id,
                        app_name,
                        data_type,
                        from_date,
                        to_date,
                        status,
                        message,
                        api_response,
                        error_details,
                        row_count,
                        created_at,
                        event_filter
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, %s
                    )
                """, (
                    query_result_id,
                    query_result_id,
                    current_user,
                    data.get('accountType'),
                    data.get('accountId'),
                    data.get('appId'),
                    data.get('appName'),
                    data.get('dataType'),
                    from_date,
                    to_date,
                    data.get('status', 'error'),
                    data.get('message', ''),
                    json.dumps(data.get('apiResponse', {})),
                    json.dumps(data.get('errorDetails', {})),
                    data.get('rowCount', 0),
                    data.get('eventFilter', None)
                ))
                logger.info(f"创建新的查询日志记录: {query_result_id}")
            else:
                # Update existing row
                update_fields = []
                update_values = []
                
                if data.get('status') is not None:
                    update_fields.append("status = %s")
                    update_values.append(data.get('status'))
                if data.get('message') is not None:
                    update_fields.append("message = %s")
                    update_values.append(data.get('message'))
                if data.get('apiResponse') is not None:
                    update_fields.append("api_response = %s")
                    update_values.append(json.dumps(data.get('apiResponse')))
                if data.get('errorDetails') is not None:
                    update_fields.append("error_details = %s")
                    update_values.append(json.dumps(data.get('errorDetails')))
                if data.get('rowCount') is not None:
                    update_fields.append("row_count = %s")
                    update_values.append(data.get('rowCount'))
                if data.get('appName') is not None:
                    update_fields.append("app_name = %s")
                    update_values.append(data.get('appName'))
                if from_date is not None:
                    update_fields.append("from_date = %s")
                    update_values.append(from_date)
                if to_date is not None:
                    update_fields.append("to_date = %s")
                    update_values.append(to_date)
                if data.get('eventFilter') is not None:
                    update_fields.append("event_filter = %s")
                    update_values.append(data.get('eventFilter'))
                
                
                if update_fields:
                    update_values.append(query_result_id)
                    cursor.execute(f"""
                        UPDATE query_logs 
                        SET {', '.join(update_fields)}
                        WHERE id = %s OR query_result_id = %s
                    """, tuple(update_values + [query_result_id]))
                    logger.info(f"更新查询日志记录: {query_result_id}")
            
            bump_home_cache_generation("query_log:update")
            return jsonify({'message': 'Query log updated successfully'})
            
    except Exception as e:
        logger.error(f"更新查询日志失败: {str(e)}", exc_info=True)
        return jsonify({'message': str(e)}), 500

@app.route('/api/query-logs', methods=['GET'])
@token_required
def get_query_logs(current_user):
    try:
        account_type = request.args.get('accountType')
        account_id = request.args.get('accountId')
        
        if not account_type or not account_id:
            return jsonify({'success': False, 'message': 'Missing Account Information'}), 400
            
        with get_db_cursor() as cursor:
            cursor.execute("""
                SELECT * FROM query_logs 
                WHERE account_type = %s AND account_id = %s
                ORDER BY created_at DESC
            """, (account_type, account_id))
            
            logs = cursor.fetchall()
        return jsonify({
            'success': True,
            'logs': [{
                'key': log['query_result_id'],
                'accountType': log['account_type'],
                'accountId': log['account_id'],
                'appId': log['app_id'],
                'dataType': log['data_type'],
                'dateRange': f"{log['from_date']} To {log['to_date']}",
                'status': log['status'],
                'message': log['message'],
                'apiResponse': json.loads(log['api_response']) if log['api_response'] else {},
                'errorDetails': json.loads(log['error_details']) if log['error_details'] else {},
                'rowCount': log['row_count'],
                'event_filter': log['event_filter']
            } for log in logs]
        })
    except Exception as e:
        logger.error(f"获取查询日志失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/query-logs/<query_result_id>', methods=['DELETE'])
@token_required
def delete_query_log(current_user, query_result_id):
    try:
        # Get account and required params
        account_type = request.args.get('accountType')
        account_id = request.args.get('accountId')
        data_type = request.args.get('dataType')
        from_date = request.args.get('fromDate')
        to_date = request.args.get('toDate')
        
        if not all([account_type, account_id, data_type, from_date, to_date]):
            return jsonify({
                'status': 'error',
                'message': 'Missing Required Parameters: accountType, accountId, dataType, fromDate, toDate'
            }), 400
            
        # Get current user ID
        current_user_id = current_user if isinstance(current_user, str) else current_user.get('id') or current_user.get('user_id')
        if not current_user_id:
            logger.error("无法获取用户ID")
            return jsonify({
                'status': 'error',
                'message': 'Failed to Get User Information'
            }), 401
            
        logger.info(f"开始删除查询日志: {query_result_id}, 账户: {account_type}/{account_id}, 数据类型: {data_type}, 日期范围: {from_date} 至 {to_date}")
        
        # DB connection via with
        with get_db_cursor() as cursor:
            # Fetch log by id or query_result_id; verify user
            cursor.execute("""
                SELECT * FROM query_logs 
                WHERE (id = %s OR query_result_id = %s)
                AND user_id = %s
            """, (query_result_id, query_result_id, current_user_id))
            
            log = cursor.fetchone()
            if not log:
                logger.warning(f"未找到要删除的查询日志: {query_result_id}")
                logger.debug(f"SQL查询参数: id={query_result_id}, user_id={current_user_id}")
                return jsonify({
                    'status': 'error',
                    'message': 'Query Log Not Found or Status Does Not Allow Deletion'
                }), 404
            
            # Get download URL from API response
            api_response = json.loads(log['api_response']) if log['api_response'] else {}
            download_url = api_response.get('downloadUrl', '')
            
            # Extract filename from download URL
            filename = download_url.split('/')[-1] if download_url else None
            
            # Delete DB row by id/query_result_id; verify user
            cursor.execute("""
                DELETE FROM query_logs 
                WHERE (id = %s OR query_result_id = %s)
                AND user_id = %s
            """, (query_result_id, query_result_id, current_user_id))
            
            # Delete physical file if download_url set
            download_url = log.get('download_url')
            if not download_url and log.get('api_response'):
                # Get downloadUrl from api_response
                try:
                    api_response = json.loads(log['api_response'])
                    download_url = api_response.get('downloadUrl', '')
                except:
                    download_url = ''
            
            if download_url:
                try:
                    # Extract filename from downloadUrl
                    if download_url.startswith('/api/download/'):
                        filename = download_url.replace('/api/download/', '')
                    elif '/' in download_url:
                        filename = download_url.split('/')[-1]
                    else:
                        filename = download_url
                    
                    logger.info(f"尝试删除文件: {filename}")
                    
                    # Candidate file paths
                    possible_paths = [
                        os.path.join('temp', filename),  # Relative path first
                        os.path.join(os.getcwd(), 'temp', filename),  # Absolute path
                        filename  # Bare filename last resort
                    ]
                    
                    file_deleted = False
                    for path in possible_paths:
                        if os.path.exists(path) and os.path.isfile(path):
                            try:
                                os.remove(path)
                                logger.info(f"成功删除物理文件: {path}")
                                file_deleted = True
                                break
                            except OSError as e:
                                logger.warning(f"删除物理文件失败 {path}: {str(e)}")
                                continue
                    
                    if not file_deleted:
                        logger.warning(f"未找到或无法删除物理文件: {filename}")
                        # List temp dir files for debug
                        try:
                            temp_dirs = ['temp', os.path.join(os.getcwd(), 'temp')]
                            for temp_dir in temp_dirs:
                                if os.path.exists(temp_dir):
                                    logger.info(f"temp目录 {temp_dir} 下的文件:")
                                    for file in os.listdir(temp_dir):
                                        if os.path.isfile(os.path.join(temp_dir, file)):
                                            logger.info(f"  - {file}")
                        except Exception as list_error:
                            logger.error(f"列出temp目录文件失败: {str(list_error)}")
                        
                except Exception as file_error:
                    logger.error(f"删除物理文件时出错: {str(file_error)}")
                    # File delete failure does not block DB delete
            else:
                logger.info("记录没有download_url，跳过物理文件删除")
            
            logger.info(f"查询日志删除成功: {query_result_id}")
            bump_home_cache_generation("query_log:delete")
            return jsonify({
                'status': 'success',
                'message': 'Query Log Deleted Successfully'
            })
            
    except Exception as e:
        logger.error(f"删除查询日志失败: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Failed to Delete Query Log: {str(e)}'
        }), 500

@app.route('/api/check-duplicate-query', methods=['POST'])
@token_required
def check_duplicate_query(current_user):
    """Check duplicate queries; supports batch."""
    try:
        data = request.get_json()
        
        # Single appId or batch appIds
        app_ids = data.get('appIds', [data.get('appId')])
        if not app_ids or not any(app_ids):
            return jsonify({
                'status': 'error',
                'message': 'Missing appId or appIds Field'
            }), 400
        
        # Drop empty values
        app_ids = [app_id for app_id in app_ids if app_id]
        
        required_fields = ['accountType', 'accountId', 'dataType', 'fromDate', 'toDate']
        for field in required_fields:
            if field not in data:
                return jsonify({
                    'status': 'error',
                    'message': f'缺少必要字段: {field}'
                }), 400

        # Get event filter param
        event_filter = data.get('eventFilter', '')
        # Get mode param
        mode = data.get('mode', 'normal')

        with get_db_cursor() as cursor:
            # Batch duplicate check; omit mode filter if column missing
            placeholders = ','.join(['%s'] * len(app_ids))
            params_with_mode = (
                current_user,
                data['accountType'],
                data['accountId'],
                data['dataType'],
                data['fromDate'],
                data['toDate'],
                *app_ids,
                event_filter,
                mode
            )
            sql_with_mode = f"""
                SELECT id, status, message, created_at, app_id
                FROM query_logs 
                WHERE user_id = %s
                AND account_type = %s 
                AND account_id = %s 
                AND data_type = %s 
                AND from_date = %s 
                AND to_date = %s 
                AND app_id IN ({placeholders})
                AND COALESCE(event_filter, '') = %s
                AND mode = %s
                ORDER BY created_at DESC
                LIMIT 1
            """
            params_without_mode = (
                current_user,
                data['accountType'],
                data['accountId'],
                data['dataType'],
                data['fromDate'],
                data['toDate'],
                *app_ids,
                event_filter
            )
            sql_without_mode = f"""
                SELECT id, status, message, created_at, app_id
                FROM query_logs 
                WHERE user_id = %s
                AND account_type = %s 
                AND account_id = %s 
                AND data_type = %s 
                AND from_date = %s 
                AND to_date = %s 
                AND app_id IN ({placeholders})
                AND COALESCE(event_filter, '') = %s
                ORDER BY created_at DESC
                LIMIT 1
            """
            existing_record = None
            try:
                cursor.execute(sql_with_mode, params_with_mode)
                existing_record = cursor.fetchone()
            except Exception as e:
                err_msg = str(e) if e else ''
                # 1054 Unknown column 'mode': fallback query without mode
                if '1054' in err_msg and 'mode' in err_msg.lower():
                    cursor.execute(sql_without_mode, params_without_mode)
                    existing_record = cursor.fetchone()
                else:
                    raise
            
            if existing_record:
                _ca = existing_record.get('created_at')
                _ca_str = _ca.strftime('%Y-%m-%dT%H:%M:%S') if _ca and hasattr(_ca, 'strftime') else (str(_ca) if _ca else '')
                return jsonify({
                    'status': 'success',
                    'isDuplicate': True,
                    'record': {
                        'id': existing_record['id'],
                        'status': existing_record['status'],
                        'message': existing_record.get('message') or '',
                        'createdAt': _ca_str,
                        'appId': existing_record.get('app_id')
                    }
                })
            
            return jsonify({
                'status': 'success',
                'isDuplicate': False
            })
            
    except Exception as e:
        logger.error(f"检查重复查询失败: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/api/download-all', methods=['GET'])
@token_required
def download_all(current_user):
    # Get mode; default normal
    mode = request.args.get('mode', 'normal')
    temp_dir = None  # Initialize temp_dir
    try:
        # Get current user ID
        current_user_id = current_user if isinstance(current_user, str) else current_user.get('id') or current_user.get('user_id')
        if not current_user_id:
            logger.error("无法获取用户ID")
            return jsonify({
                'status': 'error',
                'message': 'Failed to Get User Information'
            }), 401
        logger.info(f"开始处理用户 {current_user_id} 的下载全部请求")
        # Fetch successful logs for mode
        with get_db_cursor() as cursor:
            cursor.execute("""
                SELECT * FROM query_logs 
                WHERE user_id = %s 
                AND status = 'success'
                AND mode = %s
                ORDER BY created_at DESC
            """, (current_user_id, mode))
            records = cursor.fetchall()
            if not records:
                logger.info(f"用户 {current_user_id} 没有可下载的记录")
                return jsonify({
                    'status': 'error',
                    'message': 'No Downloadable Records'
                }), 404
            # Create temp dir (absolute path)
            temp_dir = os.path.join(TEMP_DIR, f'download_all_{current_user_id}_{int(time.time())}')
            os.makedirs(temp_dir, exist_ok=True)
            logger.info(f"创建临时目录: {temp_dir}")
            # Download all files
            downloaded_files = []
            for record in records:
                try:
                    api_response = json.loads(record['api_response'])
                    download_url = api_response.get('downloadUrl')
                    if not download_url:
                        logger.warning(f"记录 {record['id']} 没有下载URL")
                        continue
                    # Extract original filename from URL
                    original_filename = download_url.split('/')[-1]
                    # Source file path (absolute)
                    src_file_path = os.path.join(TEMP_DIR, original_filename)
                    # Regenerate filename
                    account_id_value = record['account_id']
                    app_id = record['app_id']
                    data_type = record['data_type']
                    from_date = record['from_date'].strftime('%Y-%m-%d')
                    to_date = record['to_date'].strftime('%Y-%m-%d')
                    event_filter = record.get('event_filter', '')
                    timestamp = record['created_at'].strftime('%Y%m%d_%H%M%S')
                    base_name = f"RawData({account_id_value})"
                    is_event_type = data_type in ['In-App-Event-Postbacks', 'In-App-Event-Non-Organic', 'Retargeting-In-App-Event-Postbacks', 'Retargeting-In-App-Event-Non-Organic']
                    if is_event_type:
                        if event_filter and event_filter.strip():
                            data_type_with_filter = f"{data_type}(Filter-On)"
                        else:
                            data_type_with_filter = f"{data_type}(Filter-Off)"
                    else:
                        data_type_with_filter = data_type
                    if from_date != to_date:
                        new_filename = f"{base_name}_{app_id}_{data_type_with_filter}_{from_date}——{to_date}_{timestamp}.csv"
                    else:
                        new_filename = f"{base_name}_{app_id}_{data_type_with_filter}_{from_date}_{timestamp}.csv"
                    new_file_path = os.path.join(temp_dir, new_filename)
                    if os.path.exists(src_file_path):
                        # Copy via local filesystem
                        with open(src_file_path, 'rb') as src_f, open(new_file_path, 'wb') as dst_f:
                            dst_f.write(src_f.read())
                        downloaded_files.append(new_file_path)
                        logger.info(f"成功复制文件: {src_file_path} -> {new_file_path}")
                    else:
                        logger.warning(f"文件不存在: {src_file_path}")
                except Exception as e:
                    logger.error(f"处理文件失败: {str(e)}")
                    continue
            if not downloaded_files:
                logger.warning(f"用户 {current_user_id} 没有成功下载的文件")
                return jsonify({
                    'status': 'error',
                    'message': 'No Downloadable Files'
                }), 404
            # Create ZIP
            file_count = len(downloaded_files)
            zip_filename = f'RawData_Bundle_{file_count}_Results.zip'  # ZIP name: ASCII letters, digits, underscore only
            zip_path = os.path.join(temp_dir, zip_filename)
            with zipfile.ZipFile(zip_path, 'w') as zipf:
                for file_path in downloaded_files:
                    zipf.write(file_path, os.path.basename(file_path))
            logger.info(f"成功创建ZIP文件: {zip_filename}")
            # Send ZIP file
            return send_file(
                zip_path,
                as_attachment=True,
                download_name=zip_filename,
                mimetype='application/zip',
                conditional=True  # Enable streaming and range requests
            )
    except Exception as e:
        logger.error(f"下载全部文件失败: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'下载全部文件失败: {str(e)}'
        }), 500

@app.route('/api/delete-all', methods=['DELETE'])
@token_required
def delete_all(current_user):
    # Get mode; default normal
    mode = request.args.get('mode', 'normal')
    try:
        # Get current user ID
        current_user_id = current_user if isinstance(current_user, str) else current_user.get('id') or current_user.get('user_id')
        if not current_user_id:
            logger.error("无法获取用户ID")
            return jsonify({
                'status': 'error',
                'message': 'Failed to Get User Information'
            }), 401

        logger.info(f"Starting delete all request for user {current_user_id}, mode: {mode}")
        
        # Fetch all user logs for mode
        with get_db_cursor() as cursor:
            cursor.execute("""
                SELECT id, account_type, account_id, download_url
                FROM query_logs 
                WHERE user_id = %s
                AND mode = %s
            """, (current_user_id, mode))
            
            records = cursor.fetchall()
            logger.info(f"Found {len(records)} records to delete")
            
            if not records:
                return jsonify({
                    'status': 'success',
                    'message': '没有找到需要删除的记录',
                    'deletedCount': 0
                })
            
            # Delete rows one by one; remove files
            deleted_count = 0
            for record in records:
                try:
                    # Delete DB row
                    cursor.execute("""
                        DELETE FROM query_logs 
                        WHERE `id` = %s
                    """, (record['id'],))
                    
                    if cursor.rowcount > 0:
                        deleted_count += 1
                        
                        # Delete physical file if download_url set
                        download_url = record.get('download_url')
                        if download_url:
                            try:
                                # Extract filename from downloadUrl
                                if download_url.startswith('/api/download/'):
                                    filename = download_url.replace('/api/download/', '')
                                elif '/' in download_url:
                                    filename = download_url.split('/')[-1]
                                else:
                                    filename = download_url
                                
                                logger.info(f"尝试删除文件: {filename}")
                                
                                # Candidate file paths
                                possible_paths = [
                                    os.path.join('temp', filename),  # Relative path first
                                    os.path.join(os.getcwd(), 'temp', filename),  # Absolute path
                                    filename  # Bare filename last resort
                                ]
                                
                                file_deleted = False
                                for path in possible_paths:
                                    if os.path.exists(path) and os.path.isfile(path):
                                        try:
                                            os.remove(path)
                                            logger.info(f"成功删除物理文件: {path}")
                                            file_deleted = True
                                            break
                                        except OSError as e:
                                            logger.warning(f"删除物理文件失败 {path}: {str(e)}")
                                            continue
                                
                                if not file_deleted:
                                    logger.warning(f"未找到或无法删除物理文件: {filename}")
                                    
                            except Exception as file_error:
                                logger.error(f"删除物理文件时出错: {str(file_error)}")
                                # File delete failure does not block DB delete
                        
                except Exception as record_error:
                    logger.error(f"删除记录 {record['id']} 失败: {str(record_error)}")
                    # Continue on error; do not abort batch
            
            logger.info(f"成功删除 {deleted_count} 条记录")
            if deleted_count > 0:
                bump_home_cache_generation("query_log:delete_all")
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully Deleted {deleted_count} Records',
            'deletedCount': deleted_count
        })
        
    except Exception as e:
        logger.error(f"删除全部记录失败: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to Delete All Records: {str(e)}'
        }), 500

@app.errorhandler(Exception)
def handle_error(error):
    logger.error(f"全局错误处理: {str(error)}")
    response = {
        "status": "error",
        "message": str(error),
        "error_type": error.__class__.__name__
    }
    if isinstance(error, HTTPException):
        response["status_code"] = error.code
    else:
        response["status_code"] = 500
    return jsonify(response)

@app.route('/api/download', methods=['POST'])
def download_data():
    try:
        data = request.get_json()
        app_id = data.get('app_id')
        data_type = data.get('data_type')
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        account_type = data.get('account_type')
        
        # Build API URL
        api_url = f"https://hq.appsflyer.com/export/{app_id}/{data_type}/v5"
        
        # Set request params
        params = {
            'api_token': 'YOUR_API_TOKEN',
            'from': from_date,
            'to': to_date
        }
        
        # Send request with SSL config
        try:
            response = requests.get(api_url, params=params, verify=True, timeout=30)
        except requests.exceptions.SSLError as ssl_error:
            logger.error(f'SSL连接错误: {str(ssl_error)}')
            # On SSL failure, retry without verify (fallback only)
            logger.warning('尝试禁用SSL验证作为备用方案')
            response = requests.get(api_url, params=params, verify=False, timeout=30)
        
        if response.status_code != 200:
            return jsonify({'error': 'Failed to fetch data from AppsFlyer API'}), 500
        
        # Save response to temp file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Build new file naming scheme
        # 1. Prefix: RawData
        # 2. Add account_type
        # 3. Event filter tag; no event_filter param (Filter-Off)
        base_name = f"RawData({account_type})"
        
        # Check if event-type data
        is_event_type = data_type in ['In-App-Event-Postbacks', 'In-App-Event-Non-Organic', 'Retargeting-In-App-Event-Postbacks', 'Retargeting-In-App-Event-Non-Organic']
        
        # Event filter tag; no event_filter param here (Filter-Off)
        if is_event_type:
            data_type_with_filter = f"{data_type}(Filter-Off)"
        else:
            data_type_with_filter = data_type
        
        # Build full filename
        if from_date != to_date:
            filename = f"{base_name}_{app_id}_{data_type_with_filter}_{from_date}——{to_date}_{timestamp}.csv"
        else:
            filename = f"{base_name}_{app_id}_{data_type_with_filter}_{from_date}_{timestamp}.csv"
        
        filepath = os.path.join('temp', filename)
        
        # Ensure temp dir exists
        os.makedirs('temp', exist_ok=True)
        
        # Process Postback URL fields (iOS only)
        processed_csv_content = response.text
        try:
            # Check iOS platform rows
            if 'platform' in response.text.lower() and 'ios' in response.text.lower():
                logger.info("检测到iOS平台数据，开始处理Postback URL字段")
                processed_csv_content = process_postback_urls_in_csv(response.text)
                logger.info("Postback URL字段处理完成")
        except Exception as e:
            logger.warning(f"处理Postback URL字段失败，使用原始数据: {str(e)}")
            processed_csv_content = response.text
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(processed_csv_content)
        
        logger.info(f'数据已保存到文件: {filepath}')

        # Count Is Primary Attribution for event types only
        primary_attribution_count = 0
        event_types = [
            'In-App-Event-Postbacks', 'Retargeting-In-App-Event-Postbacks',
            'In-App-Event-Non-Organic', 'Retargeting-In-App-Event-Non-Organic'
        ]
        
        if data_type in event_types:
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    # Read CSV header row
                    header_line = f.readline().strip()
                    csv_headers = header_line.split(',')
                    
                    # Find Is Primary Attribution column index
                    primary_attribution_index = None
                    for i, header in enumerate(csv_headers):
                        if header.lower() in ['is primary attribution', 'is_primary_attribution', 'isprimaryattribution']:
                            primary_attribution_index = i
                            break
                    
                    if primary_attribution_index is not None:
                        # Read rows and aggregate
                        for line in f:
                            values = line.strip().split(',')
                            if len(values) > primary_attribution_index:
                                if values[primary_attribution_index].lower() == 'true':
                                    primary_attribution_count += 1
            except Exception as e:
                logger.error(f"统计Primary Attribution数量失败: {str(e)}")
        
        # Dedupe AppsFlyer ID count via csv.reader
        afid_deduplication_count = 0
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                csv_headers = next(reader)
                afid_index = None
                for i, header in enumerate(csv_headers):
                    if header.lower().replace('_', '').replace(' ', '') == 'appsflyerid':
                        afid_index = i
                        break
                if afid_index is not None:
                    afids = set()
                    for row in reader:
                        if len(row) > afid_index:
                            afid = row[afid_index].strip().replace('\r','').replace('\n','')
                            if afid:
                                afids.add(afid)
                    logger.info(f"去重AppsFlyer ID统计: {afids}")
                    afid_deduplication_count = len(afids)
        except Exception as e:
            logger.error(f"统计AppsFlyer ID去重数量失败: {str(e)}")

        # Get app_name from DB
        app_name = None
        try:
            with get_db_cursor() as cursor:
                cursor.execute("""
                    SELECT app_name FROM accounts 
                    WHERE account_type = %s AND account_id = %s AND app_id = %s
                    LIMIT 1
                """, (account_type, account_type, app_id))
                result = cursor.fetchone()
                if result:
                    app_name = result['app_name']
        except Exception as e:
            logger.error(f"从数据库获取app_name失败: {str(e)}")

        # If no app_name in DB, read from file
        if not app_name:
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    header_line = f.readline().strip()
                    # Find app_name column
                    csv_headers = [h.strip() for h in header_line.split(',')]
                    app_name_index = None
                    for i, header in enumerate(csv_headers):
                        if header.lower() in ['app_name', 'appname', 'app name']:
                            app_name_index = i
                            break
                    if app_name_index is not None:
                        # Read first data row
                        data_line = f.readline().strip()
                        if data_line:
                            values = [v.strip() for v in data_line.split(',')]
                            # Pad row to header length
                            if len(values) < len(csv_headers):
                                values += [''] * (len(csv_headers) - len(values))
                            candidate = values[app_name_index]
                            logger.info(f"读取到 app_name 字段内容: {candidate}")
                            # Set app_name only when non-empty
                            if candidate:
                                app_name = candidate
                                # Persist app_name to DB
                                try:
                                    with get_db_cursor() as cursor:
                                        cursor.execute("""
                                            INSERT INTO accounts (
                                                id, account_type, account_name, account_id, 
                                                app_id, app_name, create_time, update_time
                                            ) VALUES (
                                                %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP
                                            ) AS new
                                            ON DUPLICATE KEY UPDATE 
                                                app_name = new.app_name,
                                                update_time = CURRENT_TIMESTAMP
                                        """, (
                                            str(uuid.uuid4()),
                                            account_type,
                                            account_type,
                                            account_type,
                                            app_id,
                                            app_name
                                        ))
                                        logger.info(f"已将app_name保存到数据库: {app_name}")
                                except Exception as e:
                                    logger.error(f"保存app_name到数据库失败: {str(e)}")
            except Exception as e:
                logger.error(f"从文件中读取app_name失败: {str(e)}")
        # Return None if app_name still empty
        if not app_name:
            # Try app_name from apps_finder
            try:
                with get_db_cursor() as cursor:
                    cursor.execute("SELECT app_name FROM apps_finder WHERE app_id = %s", (app_id,))
                    result = cursor.fetchone()
                    if result and result['app_name']:
                        app_name = result['app_name']
                        logger.info(f"从apps_finder获取到App Name: {app_name}")
            except Exception as e:
                logger.warning(f"从apps_finder查询App Name失败: {str(e)}")
        
        # Set app_name to None if still missing
        if not app_name:
            app_name = None

        # Build response payload
        response_data = {
            'status': 'success',
            'message': 'Data Fetched Successfully',
            'downloadUrl': f'/api/download/{filename}',
            'details': {
                'appId': app_id,
                'appName': app_name,
                'dataType': data_type,
                'dateRange': f'{from_date} To {to_date}',
                'rowCount': len(response.text.splitlines()) - 1,
                'afidDeduplicationCount': afid_deduplication_count,
                'config': {
                    'url': api_url,
                    'params': params,
                    'headers': {'api_token': 'YOUR_API_TOKEN'}
                }
            }
        }
        
        # Add primaryAttributionCount for event types only
        if data_type in event_types:
            response_data['details']['primaryAttributionCount'] = primary_attribution_count

        return response_data
        
    except Exception as e:
        logger.error(f'下载数据时发生错误: {str(e)}')
        return jsonify({'error': str(e)}), 500

# Safe CSV split helper
def safe_split(line, expected_len):
    # Handle trailing empty fields without comma
    # Use rsplit for field count
    values = [v.strip() for v in line.rstrip('\n').split(',')]
    if len(values) < expected_len:
        values += [''] * (expected_len - len(values))
    elif len(values) > expected_len:
        # Merge overflow fields into last column
        values = values[:expected_len-1] + [','.join(values[expected_len-1:])]
    return values

@app.route('/api/apps-finder', methods=['GET'])
def get_apps_finder():
    """
    Pagination via page and pageSize; returns { total, data }.
    Legacy: omit params to return all rows.
    """
    try:
        cache_key = None
        if _home_cache_enabled() and not _is_apps_finder_no_cache_request():
            cache_key = _apps_finder_cache_key("list", current_user={"id": "apps-finder"})
            cached_payload = _home_cache_get_json(cache_key)
            if cached_payload is not None:
                response = jsonify(cached_payload)
                response.headers["X-AppsFinder-Cache"] = "HIT"
                return response

        page = request.args.get('page', type=int)
        page_size = request.args.get('pageSize', type=int)
        params = []
        where = []
        # Support filter params
        os = request.args.get('os')
        if os:
            where.append('os=%s')
            params.append(os)
        category = request.args.get('category')
        if category:
            where.append('category=%s')
            params.append(category)
        app_id = request.args.get('appId')
        if app_id:
            where.append('app_id=%s')
            params.append(app_id)
        app_name = request.args.get('appName')
        if app_name:
            where.append('app_name LIKE %s')
            params.append(f'%{app_name}%')
            logger.debug(f"App Name搜索参数: {app_name}")
        country = request.args.get('country')
        if country:
            # country stored lowercase; frontend sends lowercase
            where.append('LOWER(country)=LOWER(%s)')
            params.append(country)
            logger.debug(f"Country筛选参数: {country}")
        # Keyword search (app_id / app_name / developer)
        keyword = request.args.get('keyword')
        if keyword:
            where.append('(app_id LIKE %s OR app_name LIKE %s OR developer LIKE %s)')
            like_kw = f'%{keyword}%'
            params.extend([like_kw, like_kw, like_kw])
            logger.debug(f"Keyword搜索参数: {keyword}")
        wheresql = ('WHERE ' + ' AND '.join(where)) if where else ''
        # Pagination
        if page and page_size:
            offset = (page - 1) * page_size
            # Count total first
            count_query = f"SELECT COUNT(*) as total FROM apps_finder {wheresql}"
            logger.debug(f"执行计数查询: {count_query} 参数: {params}")
            with get_db_cursor() as cursor:
                cursor.execute(count_query, tuple(params))
                total = cursor.fetchone()['total']
            # Fetch current page
            data_query = f"SELECT * FROM apps_finder {wheresql} ORDER BY created_at DESC LIMIT %s OFFSET %s"
            page_params = params + [page_size, offset]
            logger.debug(f"Execute data query.: {data_query} 参数: {page_params}")
            with get_db_cursor() as cursor:
                cursor.execute(data_query, tuple(page_params))
                rows = cursor.fetchall()
            result = [
                {
                    "os": row["os"],
                    "appId": row["app_id"],
                    "country": row.get("country"),
                    "category": row["category"],
                    "appName": row["app_name"],
                    "developer": row["developer"],
                    "developerUrl": row.get("developer_url"),
                    "description": row["description"],
                    "url": row["url"],
                    "iconUrl": row.get("icon_url"),
                    "rating": row.get("rating"),
                    "ratingCount": row.get("rating_count"),
                    "keywords": row.get("keywords")
                }
                for row in rows
            ]
            payload = {"total": total, "data": result}
            if cache_key:
                _home_cache_set_json(cache_key, payload, APPS_FINDER_CACHE_TTL_LIST)
            response = jsonify(payload)
            response.headers["X-AppsFinder-Cache"] = "MISS" if cache_key else "BYPASS"
            return response
        else:
            # Legacy: return full list
            query = f"SELECT * FROM apps_finder {wheresql} ORDER BY created_at DESC"
            with get_db_cursor() as cursor:
                cursor.execute(query, tuple(params))
                rows = cursor.fetchall()
            result = [
                {
                    "os": row["os"],
                    "appId": row["app_id"],
                    "country": row.get("country"),
                    "category": row["category"],
                    "appName": row["app_name"],
                    "developer": row["developer"],
                    "developerUrl": row.get("developer_url"),
                    "description": row["description"],
                    "url": row["url"],
                    "iconUrl": row.get("icon_url"),
                    "rating": row.get("rating"),
                    "ratingCount": row.get("rating_count"),
                    "keywords": row.get("keywords")
                }
                for row in rows
            ]
            payload = result
            if cache_key:
                _home_cache_set_json(cache_key, payload, APPS_FINDER_CACHE_TTL_LIST)
            response = jsonify(payload)
            response.headers["X-AppsFinder-Cache"] = "MISS" if cache_key else "BYPASS"
            return response
    except Exception as e:
        logger.error(f"获取Apps Finder数据失败: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/apps-finder/categories', methods=['GET'])
def get_all_categories():
    """Unique non-empty categories in apps_finder; supports filters."""
    try:
        cache_key = None
        if _home_cache_enabled() and not _is_apps_finder_no_cache_request():
            cache_key = _apps_finder_cache_key("categories", current_user={"id": "apps-finder"})
            cached_payload = _home_cache_get_json(cache_key)
            if cached_payload is not None:
                response = jsonify(cached_payload)
                response.headers["X-AppsFinder-Cache"] = "HIT"
                return response

        with get_db_cursor() as cursor:
            # Build WHERE clause
            where_conditions = ["category IS NOT NULL AND category != ''"]
            params = []
            
            # Support filter params
            os = request.args.get('os')
            if os:
                where_conditions.append('os = %s')
                params.append(os)
            
            app_id = request.args.get('appId')
            if app_id:
                where_conditions.append('app_id = %s')
                params.append(app_id)
            
            app_name = request.args.get('appName')
            if app_name:
                where_conditions.append('app_name = %s')
                params.append(app_name)

            country = request.args.get('country')
            if country:
                where_conditions.append('LOWER(country)=LOWER(%s)')
                params.append(country)
            
            where_clause = ' AND '.join(where_conditions)
            query = f"SELECT DISTINCT category FROM apps_finder WHERE {where_clause}"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
        categories = [row['category'] for row in rows]
        if cache_key:
            _home_cache_set_json(cache_key, categories, APPS_FINDER_CACHE_TTL_OPTIONS)
        response = jsonify(categories)
        response.headers["X-AppsFinder-Cache"] = "MISS" if cache_key else "BYPASS"
        return response
    except Exception as e:
        logger.error(f"获取所有类目失败: {str(e)}")
        return jsonify([]), 500

@app.route('/api/apps-finder/app-names', methods=['GET'])
def get_all_app_names():
    """Unique non-empty app names in apps_finder; supports filters."""
    try:
        cache_key = None
        if _home_cache_enabled() and not _is_apps_finder_no_cache_request():
            cache_key = _apps_finder_cache_key("app-names", current_user={"id": "apps-finder"})
            cached_payload = _home_cache_get_json(cache_key)
            if cached_payload is not None:
                response = jsonify(cached_payload)
                response.headers["X-AppsFinder-Cache"] = "HIT"
                return response

        with get_db_cursor() as cursor:
            # Build WHERE clause
            where_conditions = ["app_name IS NOT NULL AND app_name != ''"]
            params = []
            
            # Support filter params
            os = request.args.get('os')
            if os:
                where_conditions.append('os = %s')
                params.append(os)
            
            category = request.args.get('category')
            if category:
                where_conditions.append('category = %s')
                params.append(category)
            
            app_id = request.args.get('appId')
            if app_id:
                where_conditions.append('app_id = %s')
                params.append(app_id)

            country = request.args.get('country')
            if country:
                where_conditions.append('LOWER(country)=LOWER(%s)')
                params.append(country)
            
            where_clause = ' AND '.join(where_conditions)
            query = f"SELECT DISTINCT app_name FROM apps_finder WHERE {where_clause} ORDER BY app_name"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
        app_names = [row['app_name'] for row in rows]
        if cache_key:
            _home_cache_set_json(cache_key, app_names, APPS_FINDER_CACHE_TTL_OPTIONS)
        response = jsonify(app_names)
        response.headers["X-AppsFinder-Cache"] = "MISS" if cache_key else "BYPASS"
        return response
    except Exception as e:
        logger.error(f"获取所有应用名称失败: {str(e)}")
        return jsonify([]), 500

@app.route('/api/apps-finder/app-name/<app_id>', methods=['GET'])
def get_app_name_by_id(app_id):
    """
    Resolve app name by App ID.
    """
    try:
        cache_key = None
        if _home_cache_enabled() and not _is_apps_finder_no_cache_request():
            cache_key = _apps_finder_cache_key("app-name", current_user={"id": "apps-finder"})
            cached_payload = _home_cache_get_json(cache_key)
            if cached_payload is not None:
                response = jsonify(cached_payload)
                response.headers["X-AppsFinder-Cache"] = "HIT"
                return response

        with get_db_cursor() as cursor:
            cursor.execute("SELECT app_name FROM apps_finder WHERE app_id = %s LIMIT 1", (app_id,))
            row = cursor.fetchone()
            payload = {"appName": row['app_name']} if row and row['app_name'] else {"appName": None}
            if cache_key:
                _home_cache_set_json(cache_key, payload, APPS_FINDER_CACHE_TTL_DETAIL)
            response = jsonify(payload)
            response.headers["X-AppsFinder-Cache"] = "MISS" if cache_key else "BYPASS"
            return response
    except Exception as e:
        logger.error(f"根据AppID查询App Name失败: {str(e)}")
        return jsonify({"appName": None})

@app.route('/api/apps-finder/url/<app_id>', methods=['GET'])
def get_app_url(app_id):
    """
    Resolve app URL by App ID.
    """
    try:
        with get_db_cursor() as cursor:
            cursor.execute("SELECT url FROM apps_finder WHERE app_id = %s LIMIT 1", (app_id,))
            row = cursor.fetchone()
            if row and row['url']:
                url = row['url']
                # Validate URL format
                if url.startswith('http://') or url.startswith('https://'):
                    logger.info(f"获取应用URL: {url}")
                    return jsonify({"url": url})
                else:
                    logger.warning(f"无效的URL格式: {url}")
                    return jsonify({"error": "Invalid URL format"}), 400
            else:
                logger.warning(f"未找到AppID对应的URL: {app_id}")
                return jsonify({"error": "URL not found"}), 404
    except Exception as e:
        logger.error(f"获取应用URL失败: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/account-configs/<account_id>/validate', methods=['POST'])
@token_required
def validate_account_config(current_user, account_id):
    """Same verify as account create/update: user-management + Push API; writes three columns."""
    result = sync_account_verify_to_db(account_id)
    if result.get("error") == "Account config not found":
        return jsonify({'error': 'Account config not found'}), 404
    return jsonify(build_validate_api_payload(result))

@app.route('/api/apps-finder/download', methods=['GET'])
def download_apps_finder():
    try:
        with get_db_cursor() as cursor:
            cursor.execute("SELECT app_id, os, app_name, developer, category, description, url, icon_url, rating, rating_count, keywords FROM apps_finder")
            rows = cursor.fetchall()
        
        if not rows:
            return jsonify({'success': False, 'message': 'No Data to Export'}), 400

        # Build DataFrame-like rows
        results = []
        for row in rows:
            results.append({
                'platform': row.get('os', ''),
                'app_id': row.get('app_id', ''),
                'app_name': row.get('app_name', ''),
                'developer': row.get('developer', ''),
                'category': row.get('category', ''),
                'description': row.get('description', ''),
                'url': row.get('url', ''),
                'icon_url': row.get('icon_url', ''),
                'rating': row.get('rating', ''),
                'rating_count': row.get('rating_count', ''),
                'keywords': row.get('keywords', '')
            })

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError("无法创建工作表")
        ws.title = '应用信息'

        # Set column order
        columns = ['platform', 'app_id', 'app_name', 'developer', 'category', 'description', 'url']
        
        # Format column headers
        def format_column_name(col_name):
            words = col_name.replace('_', ' ').split()
            return ' '.join(word.capitalize() for word in words)
        
        formatted_headers = [format_column_name(col) for col in columns]
        
        # Write header row
        ws.append(formatted_headers)
        
        # Write data rows
        for result in results:
            ws.append([
                result.get('platform', ''),
                result.get('app_id', ''),
                result.get('app_name', ''),
                result.get('developer', ''),
                result.get('category', ''),
                result.get('description', ''),
                result.get('url', '')
            ])

        # Calibri font style
        calibri_font = Font(name='Calibri')
        # Calibri bold for headers
        calibri_bold_font = Font(name='Calibri', bold=True)

        # Format header row height
        header_row = ws[1]
        ws.row_dimensions[1].height = 30  # Increase header row height
        for cell in header_row:
            cell.font = calibri_bold_font  # Apply Calibri bold
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # Remove borders
            cell.border = Border(
                left=Side(style=None), 
                right=Side(style=None), 
                top=Side(style=None), 
                bottom=Side(style=None)
            )

        # Format data rows with Calibri
        for row in ws.iter_rows(min_row=2, max_row=len(results) + 1):
            for cell in row:
                # Apply Calibri font
                cell.font = calibri_font
                # Center align; wrap text
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        # Numeric format for APP ID column
        try:
            # Safe column index via .get_loc
            app_id_col_index = formatted_headers.index('App Id')
            app_id_col_letter = get_column_letter(app_id_col_index + 1)
            for cell in ws[app_id_col_letter][1:]:  # APP ID column
                if cell.value is not None and str(cell.value).isdigit():
                    cell.number_format = '0'  # Apply number format
        except ValueError:
            logger.warning("'App Id' column not found for number formatting.")
        except Exception as e:
            logger.error(f"Error formatting App Id column: {e}")

        # Auto column width; special-case Description
        for column_index, column in enumerate(ws.columns):
            column_letter = get_column_letter(column_index + 1)
            column_name = formatted_headers[column_index]  # Column names from formatted headers
            max_length = 0

            # Max content length per column
            # Max line length per column incl. newlines
            for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                cell_value = row[column_index]
                if cell_value is not None:
                    # Longest line in multiline cells
                    lines = str(cell_value).split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    max_length = max(max_length, max_line_length)
            
            # Include header length
            header_cell_value = ws.cell(row=1, column=column_index+1).value
            if header_cell_value is not None:
                max_length = max(max_length, len(str(header_cell_value)))

            # Set column width
            if column_name == 'Description':
                # Description: small base width; limited auto-grow
                # openpyxl width is in character units (~1 per char; 30 ≈ 30 chars)
                # Min width from content; cap max
                base_width = 30  # Base width
                content_width_estimate = max_length * 0.8  # Estimate width from content length
                adjusted_width = max(base_width, min(content_width_estimate, 80))  # Min 30, max 80
                # Min width fits e.g. "Description"
                adjusted_width = max(adjusted_width, len(column_name or '') * 1.2)  # Min width fits header text
                ws.column_dimensions[column_letter].width = adjusted_width

            else:
                # Other columns: auto width with min/max
                # Tune auto-width so columns are not cramped
                adjusted_width = max_length * 1.1  # Add padding
                # Clamp width
                if adjusted_width > 60: 
                    adjusted_width = 60  # Raise max width
                if adjusted_width < 10: 
                    adjusted_width = 10
                # Min width fits header
                adjusted_width = max(adjusted_width, len(column_name or '') * 1.2)
                ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        from flask import send_file
        return send_file(
            output,
            as_attachment=True,
            download_name='APP-INFO.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Apps Finder导出失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/apps-finder/download-filtered', methods=['POST'])
@token_required
def download_apps_finder_filtered(current_user):
    try:
        data = request.get_json()
        filters = data.get('filters', {})
        
        # Build query filters
        where_conditions = []
        params = []
        
        # Platform filter
        if filters.get('os'):
            where_conditions.append("os = %s")
            params.append(filters['os'])
        
        # App ID filter
        if filters.get('appId'):
            where_conditions.append("app_id LIKE %s")
            params.append(f"%{filters['appId']}%")
        
        # App name filter
        if filters.get('appName'):
            where_conditions.append("app_name LIKE %s")
            params.append(f"%{filters['appName']}%")
        
        # Category filter
        if filters.get('category'):
            where_conditions.append("category = %s")
            params.append(filters['category'])

        # Country (geo) filter
        if filters.get('country'):
            where_conditions.append("LOWER(country)=LOWER(%s)")
            params.append(filters['country'])
        
        # Build full SQL
        base_query = """
        SELECT app_id, country, os, app_name, developer, developer_url, category, 
               description, url, icon_url, rating, rating_count, keywords, 
               created_at, updated_at
        FROM apps_finder
        """
        
        if where_conditions:
            base_query += " WHERE " + " AND ".join(where_conditions)
        
        base_query += " ORDER BY created_at DESC"
        
        logger.info(f"执行筛选查询: {base_query}")
        logger.info(f"查询参数: {params}")
        
        with get_db_cursor() as cursor:
            cursor.execute(base_query, params)
            rows = cursor.fetchall()
        
        if not rows:
            return jsonify({'success': False, 'message': 'No Data Matching Filter Criteria to Export'}), 400

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError("无法创建工作表")
        ws.title = '应用信息'

        # All DB columns in order
        columns = [
            'app_id', 'country', 'os', 'app_name', 'developer', 'developer_url', 
            'category', 'description', 'url', 'icon_url', 'rating', 'rating_count', 
            'keywords', 'created_at', 'updated_at'
        ]
        
        # Format column headers
        def format_column_name(col_name):
            column_names = {
                'app_id': 'App ID',
                'country': 'Country',
                'os': 'Platform',
                'app_name': 'App Name',
                'developer': 'Developer',
                'developer_url': 'Developer URL',
                'category': 'Category',
                'description': 'Description',
                'url': 'App URL',
                'icon_url': 'Icon URL',
                'rating': 'Rating',
                'rating_count': 'Rating Count',
                'keywords': 'Keywords',
                'created_at': 'Created At',
                'updated_at': 'Updated At'
            }
            return column_names.get(col_name, col_name.replace('_', ' ').title())
        
        formatted_headers = [format_column_name(col) for col in columns]
        
        # Write header row
        ws.append(formatted_headers)
        
        # Write data rows
        for row in rows:
            ws.append([
                row.get('app_id', ''),
                row.get('country', '').upper() if row.get('country') else '',  # Uppercase country
                row.get('os', ''),
                row.get('app_name', ''),
                row.get('developer', ''),
                row.get('developer_url', ''),
                row.get('category', ''),
                row.get('description', ''),
                row.get('url', ''),
                row.get('icon_url', ''),
                row.get('rating', ''),
                row.get('rating_count', ''),
                row.get('keywords', ''),
                row.get('created_at', ''),
                row.get('updated_at', '')
            ])

        # Calibri font style
        calibri_font = Font(name='Calibri')
        calibri_bold_font = Font(name='Calibri', bold=True)

        # Format header row
        header_row = ws[1]
        ws.row_dimensions[1].height = 30
        for cell in header_row:
            cell.font = calibri_bold_font
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style=None), 
                right=Side(style=None), 
                top=Side(style=None), 
                bottom=Side(style=None)
            )

        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1):
            for cell in row:
                cell.font = calibri_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Auto column width
        for column_index, column in enumerate(ws.columns):
            column_letter = get_column_letter(column_index + 1)
            column_name = formatted_headers[column_index]
            max_length = 0

            # Max content length per column
            for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                cell_value = row[column_index]
                if cell_value is not None:
                    lines = str(cell_value).split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    max_length = max(max_length, max_line_length)
            
            # Include header in width
            header_cell_value = ws.cell(row=1, column=column_index+1).value
            if header_cell_value is not None:
                max_length = max(max_length, len(str(header_cell_value)))

            # Set column width
            if column_name in ['Description', 'Keywords']:
                base_width = 30
                content_width_estimate = max_length * 0.8
                adjusted_width = max(base_width, min(content_width_estimate, 80))
                adjusted_width = max(adjusted_width, len(column_name or '') * 1.2)
                ws.column_dimensions[column_letter].width = adjusted_width
            else:
                adjusted_width = max_length * 1.1
                if adjusted_width > 60: 
                    adjusted_width = 60
                if adjusted_width < 10: 
                    adjusted_width = 10
                adjusted_width = max(adjusted_width, len(column_name or '') * 1.2)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filter_suffix = ""
        if filters.get('os'):
            filter_suffix += f"_{filters['os']}"
        if filters.get('category'):
            filter_suffix += f"_{filters['category']}"
        if filters.get('appId'):
            filter_suffix += f"_appId_{filters['appId']}"
        if filters.get('appName'):
            filter_suffix += f"_appName_{filters['appName']}"
        
        filename = f"Apps_Finder_Export{filter_suffix}_{timestamp}.xlsx"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Apps Finder筛选导出失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/apps-finder/delete-filtered', methods=['POST'])
@token_required
def delete_apps_finder_filtered(current_user):
    try:
        data = request.get_json()
        filters = data.get('filters', {})
        
        # Build query filters
        where_conditions = []
        params = []
        
        # Platform filter
        if filters.get('os'):
            where_conditions.append("os = %s")
            params.append(filters['os'])
        
        # App ID filter
        if filters.get('appId'):
            where_conditions.append("app_id LIKE %s")
            params.append(f"%{filters['appId']}%")
        
        # App name filter
        if filters.get('appName'):
            where_conditions.append("app_name LIKE %s")
            params.append(f"%{filters['appName']}%")
        
        # Category filter
        if filters.get('category'):
            where_conditions.append("category = %s")
            params.append(filters['category'])

        # Country (geo) filter
        if filters.get('country'):
            where_conditions.append("LOWER(country)=LOWER(%s)")
            params.append(filters['country'])
        
        # Build full SQL
        base_query = "DELETE FROM apps_finder"
        
        if where_conditions:
            base_query += " WHERE " + " AND ".join(where_conditions)
        
        logger.info(f"执行删除查询: {base_query}")
        logger.info(f"删除参数: {params}")
        
        with get_db_cursor() as cursor:
            # Count rows to delete
            count_query = "SELECT COUNT(*) as count FROM apps_finder"
            if where_conditions:
                count_query += " WHERE " + " AND ".join(where_conditions)
            
            cursor.execute(count_query, params)
            count_result = cursor.fetchone()
            total_count = count_result.get('count', 0) if count_result else 0
            
            if total_count == 0:
                return jsonify({
                    'success': False, 
                    'message': 'No Data Matching Filter Criteria to Delete'
                }), 400
            
            # Run delete
            cursor.execute(base_query, params)
            deleted_count = cursor.rowcount
            
            # Commit handled by get_db_cursor() context manager
        
        logger.info(f"成功删除 {deleted_count} 条记录")
        if deleted_count > 0:
            bump_apps_finder_cache_generation("apps_finder:delete_filtered")
        
        return jsonify({
            'success': True,
            'message': f'Successfully Deleted {deleted_count} Apps',
            'deletedCount': deleted_count
        })
        
    except Exception as e:
        logger.error(f"Apps Finder筛选删除失败: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

# =====================
# Production start (recommended):
# gunicorn --worker-class eventlet -w 1 -b 0.0.0.0:5000 backend.app:app
# Local dev: python backend/app.py
# =====================

@app.route('/api/query-logs/<query_log_id>/preview', methods=['GET'])
def preview_query_log(query_log_id):
    """Preview query log data for Home page."""
    try:
        cache_key = None
        if _home_cache_enabled() and not _is_home_no_cache_request():
            cache_key = _home_cache_key("query-log-preview", current_user={"id": "preview"})
            cached_payload = _home_cache_get_json(cache_key)
            if cached_payload is not None:
                response = jsonify(cached_payload)
                response.headers["X-Home-Cache"] = "HIT"
                return response

        # Get account from request
        account_type = request.args.get('accountType')
        account_id = request.args.get('accountId')
        
        if not account_type or not account_id:
            logger.error("缺少账户信息")
            return jsonify({"error": "Missing Account Information"}), 400
            
        # Get query log from query_logs
        with get_db_cursor() as cursor:
            cursor.execute("""
                SELECT * FROM query_logs 
                WHERE id = %s AND account_type = %s AND account_id = %s
            """, (query_log_id, account_type, account_id))
            result = cursor.fetchone()
            
            if not result:
                logger.warning(f"查询日志不存在: {query_log_id}")
                return jsonify({
                    'status': 'error',
                    'message': 'Query Log Not Found'
                }), 404
            
            # Get download URL
            download_url = result.get('download_url')
            if download_url:
                logger.info(f"找到下载URL: {download_url}")
                # Extract path from download_url
                if download_url.startswith('/api/download/'):
                    filename = download_url.replace('/api/download/', '')
                    file_path = os.path.join('temp', filename)
                    logger.info(f"使用文件路径: {file_path}")
                    
                    if os.path.exists(file_path):
                        # Read first 1000 rows for preview
                        preview_data = []
                        try:
                            with open(file_path, 'r', encoding='utf-8') as f:
                                # Read header row
                                headers = f.readline().strip().split(',')
                                
                                # Drop empty header fields
                                non_empty_headers = []
                                for i, header in enumerate(headers):
                                    if header.strip():  # Keep non-empty header names
                                        non_empty_headers.append((i, header.strip()))
                                
                                # Read first 1000 rows via csv.reader
                                import csv
                                reader = csv.reader(f)
                                for i, row in enumerate(reader):
                                    if i >= 1000:  # Limit 1000 rows
                                        break
                                    # Pad row to header length
                                    while len(row) < len(headers):
                                        row.append('')
                                    
                                    # Keep non-empty fields in order
                                    filtered_row = {}
                                    for col_index, header_name in non_empty_headers:
                                        value = row[col_index] if col_index < len(row) else ''
                                        if value.strip():  # Keep non-empty values
                                            filtered_row[header_name] = value.strip()
                                    
                                    preview_data.append(filtered_row)
                                
                        except Exception as e:
                            logger.error(f"读取文件失败: {str(e)}")
                            return jsonify({
                                'status': 'error',
                                'message': f'Failed to Read File: {str(e)}'
                            }), 500
                        
                        if cache_key:
                            _home_cache_set_json(cache_key, preview_data, HOME_CACHE_TTL_PREVIEW)
                        response = jsonify(preview_data)
                        response.headers["X-Home-Cache"] = "MISS" if cache_key else "BYPASS"
                        return response
                    else:
                        logger.error(f"文件不存在: {file_path}")
                        return jsonify({
                            'status': 'error',
                            'message': '文件不存在'
                        }), 404
                else:
                    logger.error(f"无效的下载URL格式: {download_url}")
                    return jsonify({
                        'status': 'error',
                        'message': 'Invalid Download URL Format'
                    }), 400
        
        # Error if no download_url
        logger.error(f"未找到关联的下载URL")
        return jsonify({
            'status': 'error',
            'message': 'Associated File Not Found'
        }), 404
        
    except Exception as e:
        logger.error(f"预览查询日志失败: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/api/global-account-configs', methods=['GET'])
def get_global_account_configs():
    """List all account configs (no user scope)."""
    try:
        with get_db_cursor() as cursor:
            cursor.execute("""
                SELECT id, account_name, account_type, api_token, is_default, sort_order, custom_icon, updated_at 
                FROM account_configs 
                ORDER BY sort_order ASC, account_type, account_name, updated_at DESC
            """)
            configs = cursor.fetchall()
            
            # Dedupe dict; keep newest
            unique_configs = {}
            for config in configs:
                try:
                    # Ensure required fields present
                    if all(key in config for key in ['id', 'account_name', 'account_type', 'api_token', 'is_default']):
                        key = f"{config['account_name']}_{config['account_type']}"
                        if key not in unique_configs:
                            unique_configs[key] = config
                except Exception as e:
                    logger.debug(f"处理记录时出错: {str(e)}, 记录: {config}")
                    continue
                
            config_list = list(unique_configs.values())
            logger.info(f"成功获取 {len(config_list)} 条全局账户配置")
            
            # Omit api_token from response
            sanitized_configs = []
            for config in config_list:
                sanitized_config = {k: v for k, v in config.items() if k != 'api_token'}
                sanitized_configs.append(sanitized_config)
            
            return jsonify({'configs': sanitized_configs})
            
    except Exception as e:
        logger.error(f"获取全局账户配置时发生错误: {str(e)}")
        return jsonify({'message': str(e)}), 500

@app.route('/api/users', methods=['GET'])
def get_users():
    """List all users for Manager filter."""
    try:
        with get_db_cursor() as cursor:
            # List all users from users table
            cursor.execute("""
                SELECT username 
                FROM users 
                WHERE username IS NOT NULL AND username != ''
                ORDER BY username ASC
            """)
            rows = cursor.fetchall()
            
            # Extract usernames
            usernames = [row['username'] for row in rows if row['username']]
            
            logger.info(f"成功获取 {len(usernames)} 个用户")
            return jsonify({'usernames': usernames})
            
    except Exception as e:
        logger.error(f"获取用户列表时发生错误: {str(e)}")
        return jsonify({'message': str(e)}), 500

@app.route('/api/apps-finder/platforms', methods=['GET'])
def get_all_platforms():
    """Unique non-empty OS values in apps_finder; supports filters."""
    try:
        cache_key = None
        if _home_cache_enabled() and not _is_apps_finder_no_cache_request():
            cache_key = _apps_finder_cache_key("platforms", current_user={"id": "apps-finder"})
            cached_payload = _home_cache_get_json(cache_key)
            if cached_payload is not None:
                response = jsonify(cached_payload)
                response.headers["X-AppsFinder-Cache"] = "HIT"
                return response

        with get_db_cursor() as cursor:
            # Build WHERE clause
            where_conditions = ["os IS NOT NULL AND os != ''"]
            params = []
            
            # Support filter params
            category = request.args.get('category')
            if category:
                where_conditions.append('category = %s')
                params.append(category)
            
            app_id = request.args.get('appId')
            if app_id:
                where_conditions.append('app_id = %s')
                params.append(app_id)
            
            app_name = request.args.get('appName')
            if app_name:
                where_conditions.append('app_name = %s')
                params.append(app_name)

            country = request.args.get('country')
            if country:
                where_conditions.append('LOWER(country)=LOWER(%s)')
                params.append(country)
            
            where_clause = ' AND '.join(where_conditions)
            query = f"SELECT DISTINCT os FROM apps_finder WHERE {where_clause} ORDER BY os"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
        platforms = [row['os'] for row in rows]
        if cache_key:
            _home_cache_set_json(cache_key, platforms, APPS_FINDER_CACHE_TTL_OPTIONS)
        response = jsonify(platforms)
        response.headers["X-AppsFinder-Cache"] = "MISS" if cache_key else "BYPASS"
        return response
    except Exception as e:
        logger.error(f"获取所有平台失败: {str(e)}")
        return jsonify([]), 500

@app.route('/api/apps-finder/app-ids', methods=['GET'])
def get_all_app_ids():
    """Unique non-empty app_ids in apps_finder; supports filters."""
    try:
        cache_key = None
        if _home_cache_enabled() and not _is_apps_finder_no_cache_request():
            cache_key = _apps_finder_cache_key("app-ids", current_user={"id": "apps-finder"})
            cached_payload = _home_cache_get_json(cache_key)
            if cached_payload is not None:
                response = jsonify(cached_payload)
                response.headers["X-AppsFinder-Cache"] = "HIT"
                return response

        with get_db_cursor() as cursor:
            # Build WHERE clause
            where_conditions = ["app_id IS NOT NULL AND app_id != ''"]
            params = []
            
            # Support filter params
            os = request.args.get('os')
            if os:
                where_conditions.append('os = %s')
                params.append(os)
            
            category = request.args.get('category')
            if category:
                where_conditions.append('category = %s')
                params.append(category)
            
            app_name = request.args.get('appName')
            if app_name:
                where_conditions.append('app_name = %s')
                params.append(app_name)

            country = request.args.get('country')
            if country:
                where_conditions.append('LOWER(country)=LOWER(%s)')
                params.append(country)
            
            where_clause = ' AND '.join(where_conditions)
            query = f"SELECT DISTINCT app_id FROM apps_finder WHERE {where_clause} ORDER BY app_id"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
        app_ids = [row['app_id'] for row in rows]
        if cache_key:
            _home_cache_set_json(cache_key, app_ids, APPS_FINDER_CACHE_TTL_OPTIONS)
        response = jsonify(app_ids)
        response.headers["X-AppsFinder-Cache"] = "MISS" if cache_key else "BYPASS"
        return response
    except Exception as e:
        logger.error(f"获取所有应用ID失败: {str(e)}")
        return jsonify([]), 500

@app.route('/api/apps-finder/countries', methods=['GET'])
def get_all_countries():
    """Unique non-empty countries (Country Code on cards); cascade filters."""
    try:
        cache_key = None
        if _home_cache_enabled() and not _is_apps_finder_no_cache_request():
            cache_key = _apps_finder_cache_key("countries", current_user={"id": "apps-finder"})
            cached_payload = _home_cache_get_json(cache_key)
            if cached_payload is not None:
                response = jsonify(cached_payload)
                response.headers["X-AppsFinder-Cache"] = "HIT"
                return response

        with get_db_cursor() as cursor:
            # Build WHERE clause
            where_conditions = ["country IS NOT NULL AND country != ''"]
            params = []

            # Cascade other filter params
            os = request.args.get('os')
            if os:
                where_conditions.append('os = %s')
                params.append(os)

            category = request.args.get('category')
            if category:
                where_conditions.append('category = %s')
                params.append(category)

            app_id = request.args.get('appId')
            if app_id:
                where_conditions.append('app_id = %s')
                params.append(app_id)

            app_name = request.args.get('appName')
            if app_name:
                where_conditions.append('app_name = %s')
                params.append(app_name)

            where_clause = ' AND '.join(where_conditions)
            # Return lowercase; ORDER BY alphabetically
            query = f"SELECT DISTINCT LOWER(country) AS country FROM apps_finder WHERE {where_clause} ORDER BY country"

            cursor.execute(query, params)
            rows = cursor.fetchall()
        countries = [row['country'] for row in rows if row.get('country')]
        if cache_key:
            _home_cache_set_json(cache_key, countries, APPS_FINDER_CACHE_TTL_OPTIONS)
        response = jsonify(countries)
        response.headers["X-AppsFinder-Cache"] = "MISS" if cache_key else "BYPASS"
        return response
    except Exception as e:
        logger.error(f"获取所有国家(country)失败: {str(e)}")
        return jsonify([]), 500

@app.route('/api/apps-finder/store', methods=['POST'])
def store_app_info():
    """Store or update app info; supports new fields."""
    try:
        data = request.get_json()
        print(f"接收到的数据: {data}")  # Debug info
        print(f"rating 值: {data.get('rating')}, 类型: {type(data.get('rating'))}")  # Debug info
        if not data or 'app_id' not in data:
            return jsonify({'success': False, 'message': 'Missing Required Field: app_id'}), 400
        
        app_id = data['app_id']
        country = data.get('country')
        os = data.get('os')
        app_name = data.get('app_name')
        developer = data.get('developer')
        developer_url = data.get('developer_url')
        category = data.get('category')
        description = data.get('description')
        url = data.get('url')
        icon_url = data.get('icon_url')
        rating = data.get('rating')
        rating_count = data.get('rating_count')
        keywords = data.get('keywords')
        
        # Validate required fields
        if not all([os, app_name, developer, category]):
            return jsonify({'success': False, 'message': 'Missing Required Fields: os, app_name, developer, category'}), 400
        
        with get_db_cursor() as cursor:
            # Exists check: app_id + country
            cursor.execute("SELECT app_id FROM apps_finder WHERE app_id = %s AND country = %s", (app_id, country))
            exists = cursor.fetchone()
            
            if exists:
                # If exists, skip write; return exists flag
                return jsonify({
                    'success': True,
                    'already_exists': True,
                    'message': 'App information already exists in database, skipped',
                    'app_id': app_id
                })
            else:
                # Insert new row
                sql = """
                    INSERT INTO apps_finder (
                        app_id, country, os, app_name, developer, developer_url, category, description, url,
                        icon_url, rating, rating_count, keywords
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(sql, (
                    app_id, country, os, app_name, developer, developer_url, category, description, url,
                    icon_url, rating, rating_count, keywords
                ))
                message = 'App Information Stored Successfully'
        
        bump_apps_finder_cache_generation("apps_finder:store")
        return jsonify({
            'success': True,
            'already_exists': False,
            'message': message,
            'app_id': app_id
        })
        
    except Exception as e:
        logger.error(f"存储App信息失败: {str(e)}")
        return jsonify({'success': False, 'message': f'Failed to Store: {str(e)}'}), 500


# Apps Finder upload column -> DB field map (export, snake_case, aliases)
APPS_FINDER_UPLOAD_HEADER_MAP = {
    'app id': 'app_id', 'app_id': 'app_id', 'appid': 'app_id',
    'package': 'app_id', 'package id': 'app_id', 'packageid': 'app_id',
    'bundle id': 'app_id', 'bundleid': 'app_id', 'application id': 'app_id',
    'country': 'country', 'country code': 'country',
    'platform': 'os', 'os': 'os', 'platform type': 'os',
    'app name': 'app_name', 'app_name': 'app_name', 'appname': 'app_name',
    'name': 'app_name', 'title': 'app_name', 'application name': 'app_name',
    'developer': 'developer', 'developer name': 'developer', 'vendor': 'developer',
    'developer url': 'developer_url', 'developer_url': 'developer_url', 'developerurl': 'developer_url',
    'developer website': 'developer_url', 'dev url': 'developer_url', 'developer website url': 'developer_url',
    'category': 'category', 'genre': 'category', 'categories': 'category',
    'description': 'description', 'desc': 'description', 'summary': 'description',
    'app url': 'url', 'url': 'url', 'appurl': 'url',
    'store url': 'url', 'store link': 'url', 'link': 'url', 'app link': 'url',
    'icon url': 'icon_url', 'icon_url': 'icon_url', 'iconurl': 'icon_url',
    'icon': 'icon_url', 'icon link': 'icon_url', 'artwork': 'icon_url',
    'rating': 'rating', 'score': 'rating', 'average rating': 'rating',
    'rating count': 'rating_count', 'rating_count': 'rating_count', 'ratingcount': 'rating_count',
    'reviews': 'rating_count', 'review count': 'rating_count', 'number of ratings': 'rating_count',
    'keywords': 'keywords', 'keyword': 'keywords', 'tags': 'keywords',
    'created at': 'created_at', 'created_at': 'created_at', 'createdat': 'created_at',
    'updated at': 'updated_at', 'updated_at': 'updated_at', 'updatedat': 'updated_at',
}

# Non-time fields: row must be all non-empty to insert
APPS_FINDER_NON_TIME_FIELDS = [
    'app_id', 'country', 'os', 'app_name', 'developer', 'developer_url',
    'category', 'description', 'url', 'icon_url', 'rating', 'rating_count', 'keywords'
]

# Upload header must include all non-time DB fields
APPS_FINDER_REQUIRED_HEADERS = set(APPS_FINDER_NON_TIME_FIELDS)

APPS_FINDER_UPLOAD_MAX_BYTES = 10 * 1024 * 1024  # 10MB


def _sheet_headers_match(canonical_headers):
    """Header has all required non-time fields; only then eligible for upload."""
    present = {c for c in canonical_headers if c is not None}
    return APPS_FINDER_REQUIRED_HEADERS <= present


def _normalize_header(h):
    if not h:
        return None
    s = str(h).strip().lower().replace('\ufeff', '')
    s = re.sub(r'\s+', ' ', s)  # Collapse extra spaces for header match
    return s if s else None


def _parse_apps_finder_headers(raw_headers):
    """Map first row to canonical fields; unknown -> None."""
    canonical = []
    for h in raw_headers:
        key = _normalize_header(h)
        canonical.append(APPS_FINDER_UPLOAD_HEADER_MAP.get(key) if key else None)
    return canonical


def _row_to_apps_finder_record(canonical_headers, row_values, upload_time):
    """Row to apps_finder dict; empty -> None."""
    record = {}
    for i, field in enumerate(canonical_headers):
        if field is None:
            continue
        raw = row_values[i] if i < len(row_values) else None
        if raw is not None and hasattr(raw, 'strip'):
            raw = raw.strip() if isinstance(raw, str) else raw
        val = None
        if raw is not None and str(raw).strip() != '':
            val = raw.strip() if isinstance(raw, str) else raw
        record[field] = val
    if 'country' in record and record['country']:
        record['country'] = str(record['country']).upper()[:2]
    if not record.get('created_at') and 'created_at' not in record:
        record['created_at'] = upload_time
    if not record.get('updated_at') and 'updated_at' not in record:
        record['updated_at'] = upload_time
    return record


def _is_row_valid_for_insert(record):
    """True only when all non-time fields are non-empty."""
    for f in APPS_FINDER_NON_TIME_FIELDS:
        v = record.get(f)
        if v is None or (isinstance(v, str) and v.strip() == ''):
            return False
    return True


def _parse_datetime_for_apps_finder(s):
    """Parse datetime string; None on failure."""
    if s is None or (isinstance(s, str) and not s.strip()):
        return None
    s = s.strip() if isinstance(s, str) else str(s)
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d'):
        try:
            return datetime.strptime(s[:26], fmt)
        except (ValueError, TypeError):
            continue
    return None


def _apps_finder_preview_sheets(file_storage):
    """
    Parse upload file; per-sheet validation for preview/page pick.
    Returns: (file_type: 'csv'|'xlsx', sheets: [{ name, index, valid, missing?, row_count }], error_message?)
    """
    if not file_storage or not file_storage.filename:
        return None, [], 'No file provided'
    ext = (file_storage.filename or '').lower()
    if not (ext.endswith('.csv') or ext.endswith('.xlsx')):
        return None, [], 'Invalid file type. Only CSV and XLSX are allowed.'
    file_storage.seek(0)
    size = 0
    try:
        file_storage.seek(0, 2)
        size = file_storage.tell()
        file_storage.seek(0)
    except Exception:
        pass
    if size > APPS_FINDER_UPLOAD_MAX_BYTES:
        return None, [], 'File too large. Maximum size is 10MB.'
    sheets_out = []
    if ext.endswith('.csv'):
        try:
            content = file_storage.read().decode('utf-8', errors='replace')
        except Exception:
            return None, [], 'File must be UTF-8 encoded.'
        file_storage.seek(0)
        reader = csv.reader(io.StringIO(content))
        raw_rows = list(reader)
        if not raw_rows:
            return 'csv', [], None
        raw_headers = raw_rows[0]
        canonical_headers = _parse_apps_finder_headers(raw_headers)
        present = {c for c in canonical_headers if c is not None}
        missing = sorted(APPS_FINDER_REQUIRED_HEADERS - present)
        valid = _sheet_headers_match(canonical_headers)
        sheets_out.append({
            'name': 'CSV',
            'index': 0,
            'valid': valid,
            'row_count': max(0, len(raw_rows) - 1),
            'missing': missing if missing else None
        })
        return 'csv', sheets_out, None
    else:
        try:
            # read_only=False for multi-sheet; read_only loses rows on sequential read
            wb = openpyxl.load_workbook(BytesIO(file_storage.read()), read_only=False, data_only=True)
            file_storage.seek(0)
            for idx, sheetname in enumerate(wb.sheetnames):
                ws = wb[sheetname]
                raw_rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
                if not raw_rows:
                    sheets_out.append({'name': sheetname, 'index': idx, 'valid': False, 'row_count': 0, 'missing': list(APPS_FINDER_REQUIRED_HEADERS)})
                    continue
                raw_headers = [str(c) if c is not None else '' for c in raw_rows[0]]
                canonical_headers = _parse_apps_finder_headers(raw_headers)
                present = {c for c in canonical_headers if c is not None}
                missing = sorted(APPS_FINDER_REQUIRED_HEADERS - present)
                valid = _sheet_headers_match(canonical_headers)
                sheets_out.append({
                    'name': sheetname,
                    'index': idx,
                    'valid': valid,
                    'row_count': max(0, len(raw_rows) - 1),
                    'missing': missing if missing else None
                })
            wb.close()
        except Exception as e:
            logger.warning(f"upload-preview openpyxl failed: {e}")
            return None, [], f'Invalid Excel file: {str(e)}'
        return 'xlsx', sheets_out, None


@app.route('/api/apps-finder/upload-preview', methods=['POST', 'OPTIONS'])
def upload_apps_finder_preview_route():
    """Preview upload: sheets and header match; frontend picks page before upload."""
    if request.method == 'OPTIONS':
        resp = jsonify({'status': 'ok'})
        if os.getenv('IS_LOCAL', 'false').lower() == 'true':
            origin = request.headers.get('Origin', '')
            if origin in ('http://localhost:3000', 'http://127.0.0.1:3000'):
                resp.headers.add('Access-Control-Allow-Origin', origin)
            else:
                resp.headers.add('Access-Control-Allow-Origin', 'http://localhost:3000')
        else:
            resp.headers.add('Access-Control-Allow-Origin', os.getenv('CORS_ORIGIN', 'http://localhost:3000'))
        resp.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        resp.headers.add('Access-Control-Allow-Methods', 'POST,OPTIONS')
        resp.headers.add('Access-Control-Allow-Credentials', 'true')
        return resp
    auth_header = request.headers.get('Authorization')
    token = None
    if auth_header and auth_header.startswith('Bearer '):
        token = auth_header.split(' ', 1)[1]
    elif auth_header:
        token = auth_header
    if not token:
        return jsonify({'success': False, 'message': 'Token is missing'}), 401
    try:
        data = jwt.decode(token, JWT_SECRET_KEY, algorithms=['HS256'])
        if not data.get('id'):
            return jsonify({'success': False, 'message': 'Invalid token'}), 401
    except (jwt.ExpiredSignatureError, jwt.InvalidTokenError, Exception):
        return jsonify({'success': False, 'message': 'Invalid or expired token'}), 401
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'}), 400
    f = request.files['file']
    file_type, sheets, err = _apps_finder_preview_sheets(f)
    if err:
        return jsonify({'success': False, 'message': err}), 400
    return jsonify({
        'success': True,
        'file_type': file_type,
        'sheets': sheets,
        'file_name': f.filename
    })


@app.route('/api/apps-finder/upload', methods=['POST', 'OPTIONS'])
def upload_apps_finder_route():
    """Upload: OPTIONS for CORS; POST requires auth. Single route avoids 405."""
    logger.info("apps-finder upload route: method=%s", request.method)
    if request.method == 'OPTIONS':
        resp = jsonify({'status': 'ok'})
        if os.getenv('IS_LOCAL', 'false').lower() == 'true':
            origin = request.headers.get('Origin', '')
            if origin in ('http://localhost:3000', 'http://127.0.0.1:3000'):
                resp.headers.add('Access-Control-Allow-Origin', origin)
            else:
                resp.headers.add('Access-Control-Allow-Origin', 'http://localhost:3000')
        else:
            resp.headers.add('Access-Control-Allow-Origin', os.getenv('CORS_ORIGIN', 'http://localhost:3000'))
        resp.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        resp.headers.add('Access-Control-Allow-Methods', 'POST,OPTIONS')
        resp.headers.add('Access-Control-Allow-Credentials', 'true')
        return resp
    # POST: auth token then upload
    auth_header = request.headers.get('Authorization')
    token = None
    if auth_header and auth_header.startswith('Bearer '):
        token = auth_header.split(' ', 1)[1]
    elif auth_header:
        token = auth_header
    if not token:
        return jsonify({'success': False, 'message': 'Token is missing'}), 401
    try:
        data = jwt.decode(token, JWT_SECRET_KEY, algorithms=['HS256'])
        current_user = data.get('id')
        if not current_user:
            return jsonify({'success': False, 'message': 'Invalid token'}), 401
    except jwt.ExpiredSignatureError:
        logger.warning("apps-finder upload: token expired")
        return jsonify({'success': False, 'message': 'Token has expired. Please sign in again.'}), 401
    except jwt.InvalidTokenError as e:
        logger.warning(f"apps-finder upload: invalid token: {e}")
        return jsonify({'success': False, 'message': 'Invalid token'}), 401
    except Exception as e:
        logger.warning(f"apps-finder upload token error: {e}")
        return jsonify({'success': False, 'message': 'Invalid or expired token'}), 401
    return _upload_apps_finder_impl(current_user)


def _upload_apps_finder_impl(current_user):
    """Upload CSV/XLSX to apps_finder; missing times use upload time. Optional sheet_name/sheet_index."""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        f = request.files['file']
        if not f or not f.filename:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        ext = (f.filename or '').lower()
        if not (ext.endswith('.csv') or ext.endswith('.xlsx')):
            return jsonify({'success': False, 'message': 'Invalid file type. Only CSV and XLSX are allowed.'}), 400
        f.seek(0, 2)
        size = f.tell()
        f.seek(0)
        if size > APPS_FINDER_UPLOAD_MAX_BYTES:
            return jsonify({'success': False, 'message': 'File too large. Maximum size is 10MB.'}), 400

        sheet_name = request.form.get('sheet_name', '').strip() or None
        sheet_index_val = request.form.get('sheet_index')
        sheet_index = None
        if sheet_index_val is not None and str(sheet_index_val).strip() != '':
            try:
                sheet_index = int(sheet_index_val)
                if sheet_index < 0 or sheet_index > 9999:
                    sheet_index = None
            except (ValueError, TypeError):
                sheet_index = None

        upload_time = datetime.utcnow()
        rows_data = []  # list of dict (canonical field -> value)

        if f.filename.lower().endswith('.csv'):
            try:
                content = f.read().decode('utf-8', errors='replace')
            except Exception:
                return jsonify({'success': False, 'message': 'File must be UTF-8 encoded.'}), 400
            reader = csv.reader(io.StringIO(content))
            raw_rows = list(reader)
            if not raw_rows:
                return jsonify({'success': False, 'message': 'File is empty or has no data rows.'}), 400
            raw_headers = raw_rows[0]
            canonical_headers = _parse_apps_finder_headers(raw_headers)
            if not _sheet_headers_match(canonical_headers):
                missing = APPS_FINDER_REQUIRED_HEADERS - {c for c in canonical_headers if c is not None}
                return jsonify({
                    'success': False,
                    'message': 'Headers do not match. Required columns (all non-time): ' + ', '.join(sorted(missing))
                }), 400
            for raw_row in raw_rows[1:]:
                record = _row_to_apps_finder_record(canonical_headers, raw_row, upload_time)
                if record:
                    rows_data.append(record)
        else:
            try:
                # read_only=False for correct row reads (single/multi sheet)
                wb = openpyxl.load_workbook(BytesIO(f.read()), read_only=False, data_only=True)
                best_sheet_missing = None
                for idx, sheetname in enumerate(wb.sheetnames):
                    if sheet_name is not None and sheetname != sheet_name:
                        continue
                    if sheet_index is not None and idx != sheet_index:
                        continue
                    ws = wb[sheetname]
                    raw_rows = list(ws.iter_rows(values_only=True))
                    if not raw_rows:
                        continue
                    raw_headers = [str(c) if c is not None else '' for c in raw_rows[0]]
                    canonical_headers = _parse_apps_finder_headers(raw_headers)
                    present = {c for c in canonical_headers if c is not None}
                    missing = APPS_FINDER_REQUIRED_HEADERS - present
                    if best_sheet_missing is None or len(missing) < len(best_sheet_missing[1]):
                        best_sheet_missing = (sheetname, missing, list(present))
                    if not _sheet_headers_match(canonical_headers):
                        continue
                    for raw_row in raw_rows[1:]:
                        row_values = [c if c is not None else '' for c in raw_row]
                        record = _row_to_apps_finder_record(canonical_headers, row_values, upload_time)
                        if record:
                            rows_data.append(record)
                wb.close()
            except Exception as e:
                logger.warning(f"openpyxl load failed: {e}")
                return jsonify({'success': False, 'message': f'Invalid Excel file: {str(e)}'}), 400
            if not rows_data:
                msg = 'No sheet has matching headers. Each sheet must have all non-time columns: app_id, country, os, app_name, developer, developer_url, category, description, url, icon_url, rating, rating_count, keywords.'
                if best_sheet_missing:
                    sheetname, missing, _ = best_sheet_missing
                    if missing:
                        msg += f' Closest sheet "{sheetname}" is missing: ' + ', '.join(sorted(missing)) + '.'
                msg += ' Use headers like: App ID, Country, Platform, App Name, Developer, Developer URL, Category, Description, App URL, Icon URL, Rating, Rating Count, Keywords (or snake_case).'
                return jsonify({'success': False, 'message': msg}), 400

        inserted = 0
        skipped_empty = 0
        duplicate = 0
        for record in rows_data:
            if not _is_row_valid_for_insert(record):
                skipped_empty += 1
                continue
            created_at = record.get('created_at')
            if isinstance(created_at, str):
                created_at = _parse_datetime_for_apps_finder(created_at)
            if created_at is None:
                created_at = upload_time
            updated_at = record.get('updated_at')
            if isinstance(updated_at, str):
                updated_at = _parse_datetime_for_apps_finder(updated_at)
            if updated_at is None:
                updated_at = upload_time
            def _str_strip(x):
                return str(x).strip() if x is not None else ''

            app_id = _str_strip(record.get('app_id'))[:128]
            country = _str_strip(record.get('country'))[:2].upper()
            if not app_id or not country:
                skipped_empty += 1
                continue
            with get_db_cursor() as cursor:
                cursor.execute(
                    "SELECT 1 FROM apps_finder WHERE app_id = %s AND country = %s LIMIT 1",
                    (app_id, country)
                )
                if cursor.fetchone():
                    duplicate += 1
                    continue
                try:
                    rating_val = record.get('rating')
                    if rating_val is not None and str(rating_val).strip() != '':
                        rating_val = float(rating_val)
                    else:
                        rating_val = None
                except (ValueError, TypeError):
                    rating_val = None
                try:
                    rc_val = record.get('rating_count')
                    if rc_val is not None and str(rc_val).strip() != '':
                        rc_val = int(float(rc_val))
                    else:
                        rc_val = None
                except (ValueError, TypeError):
                    rc_val = None
                os_val = _str_strip(record.get('os'))[:64]
                app_name_val = _str_strip(record.get('app_name'))[:255]
                developer_val = _str_strip(record.get('developer'))[:255]
                developer_url_val = _str_strip(record.get('developer_url'))[:512] or None
                category_val = _str_strip(record.get('category'))[:128]
                description_val = _str_strip(record.get('description')) or None
                url_val = _str_strip(record.get('url'))[:512] or None
                icon_url_val = _str_strip(record.get('icon_url'))[:512] or None
                keywords_val = _str_strip(record.get('keywords')) or None
                cursor.execute("""
                    INSERT INTO apps_finder (
                        app_id, country, os, app_name, developer, developer_url, category, description, url,
                        icon_url, rating, rating_count, keywords, created_at, updated_at
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    app_id, country,
                    os_val, app_name_val, developer_val, developer_url_val,
                    category_val, description_val, url_val, icon_url_val,
                    rating_val, rc_val, keywords_val,
                    created_at, updated_at
                ))
                inserted += 1

        msg = f'Upload completed: {inserted} row(s) inserted.'
        if skipped_empty:
            msg += f' {skipped_empty} row(s) skipped (empty required fields).'
        if duplicate:
            msg += f' {duplicate} row(s) already existed.'
        if inserted > 0:
            bump_apps_finder_cache_generation("apps_finder:upload")
        return jsonify({
            'success': True,
            'message': msg,
            'inserted_count': inserted,
            'skipped_empty_count': skipped_empty,
            'duplicate_count': duplicate
        })
    except Exception as e:
        logger.exception("apps_finder upload failed")
        return jsonify({'success': False, 'message': f'Upload failed: {str(e)}'}), 500


@app.route('/api/apps-finder/detail/<app_id>', methods=['GET'])
def get_app_info(app_id):
    """Full app info by App ID."""
    try:
        cache_key = None
        if _home_cache_enabled() and not _is_apps_finder_no_cache_request():
            cache_key = _apps_finder_cache_key("detail", current_user={"id": "apps-finder"})
            cached_payload = _home_cache_get_json(cache_key)
            if cached_payload is not None:
                response = jsonify(cached_payload)
                response.headers["X-AppsFinder-Cache"] = "HIT"
                return response

        with get_db_cursor() as cursor:
            cursor.execute("""
                SELECT app_id, os, app_name, developer, category, description, url,
                       icon_url, rating, rating_count, keywords, created_at, updated_at
                FROM apps_finder WHERE app_id = %s
            """, (app_id,))
            row = cursor.fetchone()
            
            if not row:
                return jsonify({'success': False, 'message': 'App Not Found'}), 404
            
            result = {
                'appId': row['app_id'],
                'os': row['os'],
                'appName': row['app_name'],
                'developer': row['developer'],
                'category': row['category'],
                'description': row['description'],
                'url': row['url'],
                'iconUrl': row['icon_url'],
                'rating': row['rating'],
                'ratingCount': row['rating_count'],
                'keywords': row['keywords'],
                'createdAt': row['created_at'].isoformat() if row['created_at'] else None,
                'updatedAt': row['updated_at'].isoformat() if row['updated_at'] else None
            }
            
            payload = {'success': True, 'data': result}
            if cache_key:
                _home_cache_set_json(cache_key, payload, APPS_FINDER_CACHE_TTL_DETAIL)
            response = jsonify(payload)
            response.headers["X-AppsFinder-Cache"] = "MISS" if cache_key else "BYPASS"
            return response
            
    except Exception as e:
        logger.error(f"获取应用信息失败: {str(e)}")
        return jsonify({'success': False, 'message': f'Failed to Get: {str(e)}'}), 500

@app.route('/api/apps-finder/search/<app_id>', methods=['GET'])
def search_app_by_id(app_id):
    """Search apps by App ID for typeahead."""
    try:
        cache_key = None
        if _home_cache_enabled() and not _is_apps_finder_no_cache_request():
            cache_key = _apps_finder_cache_key("search", current_user={"id": "apps-finder"})
            cached_payload = _home_cache_get_json(cache_key)
            if cached_payload is not None:
                response = jsonify(cached_payload)
                response.headers["X-AppsFinder-Cache"] = "HIT"
                return response

        with get_db_cursor() as cursor:
            # Fuzzy search by app_id
            cursor.execute("""
                SELECT app_id, os, app_name, developer, category, description, url,
                       icon_url, rating, rating_count, keywords
                FROM apps_finder 
                WHERE app_id LIKE %s 
                ORDER BY 
                    CASE WHEN app_id = %s THEN 1 ELSE 2 END,
                    app_name
                LIMIT 5
            """, (f'%{app_id}%', app_id))
            rows = cursor.fetchall()
            
            if not rows:
                payload = {'success': True, 'data': []}
                if cache_key:
                    _home_cache_set_json(cache_key, payload, APPS_FINDER_CACHE_TTL_DETAIL)
                response = jsonify(payload)
                response.headers["X-AppsFinder-Cache"] = "MISS" if cache_key else "BYPASS"
                return response
            
            results = []
            for row in rows:
                results.append({
                    'appId': row['app_id'],
                    'os': row['os'],
                    'appName': row['app_name'],
                    'developer': row['developer'],
                    'category': row['category'],
                    'description': row['description'],
                    'url': row['url'],
                    'iconUrl': row['icon_url'],
                    'rating': row['rating'],
                    'ratingCount': row['rating_count'],
                    'keywords': row['keywords']
                })
            
            payload = {'success': True, 'data': results}
            if cache_key:
                _home_cache_set_json(cache_key, payload, APPS_FINDER_CACHE_TTL_DETAIL)
            response = jsonify(payload)
            response.headers["X-AppsFinder-Cache"] = "MISS" if cache_key else "BYPASS"
            return response
            
    except Exception as e:
        logger.error(f"搜索应用信息失败: {str(e)}")
        return jsonify({'success': False, 'message': f'Failed to Search: {str(e)}'}), 500

@app.route('/api/apps-finder/delete/<app_id>', methods=['DELETE'])
def delete_app_info(app_id):
    """Delete app by App ID and country."""
    try:
        # Get country from query params
        country = request.args.get('country')
        
        with get_db_cursor() as cursor:
            # Build query filters
            if country:
                # Delete by app ID + country when country given
                cursor.execute("SELECT app_id FROM apps_finder WHERE app_id = %s AND country = %s", (app_id, country))
                exists = cursor.fetchone()
                
                if not exists:
                    return jsonify({'success': False, 'message': 'App Not Found for Specified Country'}), 404
                
                # Delete rows for country
                cursor.execute("DELETE FROM apps_finder WHERE app_id = %s AND country = %s", (app_id, country))
            else:
                # No country: delete all rows for app ID (legacy)
                cursor.execute("SELECT app_id FROM apps_finder WHERE app_id = %s", (app_id,))
                exists = cursor.fetchone()
                
                if not exists:
                    return jsonify({'success': False, 'message': 'App Not Found'}), 404
                
                # Delete all rows for app ID
                cursor.execute("DELETE FROM apps_finder WHERE app_id = %s", (app_id,))
            
            bump_apps_finder_cache_generation("apps_finder:delete")
            return jsonify({
                'success': True,
                'message': 'App Information Deleted Successfully',
                'app_id': app_id,
                'country': country
            })
            
    except Exception as e:
        logger.error(f"删除应用信息失败: {str(e)}")
        return jsonify({'success': False, 'message': f'Failed to Delete: {str(e)}'}), 500

@app.route('/api/decode-postback-url', methods=['POST'])
@token_required
def decode_postback_url(current_user):
    """Decode Postback URL and extract time fields."""
    try:
        data = request.get_json()
        postback_url = data.get('postback_url', '')
        
        if not postback_url:
            return jsonify({'error': 'Postback URL不能为空'}), 400
        
        # Parse URL params
        parsed_url = urllib.parse.urlparse(postback_url)
        query_params = urllib.parse.parse_qs(parsed_url.query)
        
        # Extract time fields
        result = {
            'attributed_touch_time': '',
            'install_time': '',
            'event_time': ''
        }
        
        # Decode CLICK_TIMESTAMP (Attributed Touch Time)
        if 'CLICK_TIMESTAMP' in query_params:
            click_timestamp = query_params['CLICK_TIMESTAMP'][0]
            decoded_click = urllib.parse.unquote(click_timestamp)
            result['attributed_touch_time'] = decoded_click
        
        # Decode INSTALL_TIMESTAMP (Install Time)
        if 'INSTALL_TIMESTAMP' in query_params:
            install_timestamp = query_params['INSTALL_TIMESTAMP'][0]
            decoded_install = urllib.parse.unquote(install_timestamp)
            result['install_time'] = decoded_install
        
        # Decode TIMESTAMP (Event Time)
        if 'TIMESTAMP' in query_params:
            timestamp = query_params['TIMESTAMP'][0]
            decoded_timestamp = urllib.parse.unquote(timestamp)
            result['event_time'] = decoded_timestamp
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        logger.error(f"解码Postback URL失败: {str(e)}")
        return jsonify({'error': f'解码失败：{str(e)}'}), 500

@app.route('/api/process-csv-postback-urls', methods=['POST'])
@token_required
def process_csv_postback_urls(current_user):
    """Batch-process Postback URL fields in CSV."""
    try:
        data = request.get_json()
        csv_data = data.get('csv_data', [])
        
        if not csv_data:
            return jsonify({'error': 'CSV数据不能为空'}), 400
        
        processed_data = []
        
        for row in csv_data:
            processed_row = row.copy()
            
            # Find Postback URL column
            postback_url_field = None
            for key, value in row.items():
                if 'postback' in key.lower() and 'url' in key.lower() and value:
                    postback_url_field = value
                    break
            
            if postback_url_field:
                try:
                    # Parse URL params
                    parsed_url = urllib.parse.urlparse(postback_url_field)
                    query_params = urllib.parse.parse_qs(parsed_url.query)
                    
                    # Extract and decode time fields
                    if 'CLICK_TIMESTAMP' in query_params:
                        click_timestamp = query_params['CLICK_TIMESTAMP'][0]
                        decoded_click = urllib.parse.unquote(click_timestamp)
                        processed_row['Attributed Touch Time'] = decoded_click
                    
                    if 'INSTALL_TIMESTAMP' in query_params:
                        install_timestamp = query_params['INSTALL_TIMESTAMP'][0]
                        decoded_install = urllib.parse.unquote(install_timestamp)
                        processed_row['Install Time'] = decoded_install
                    
                    if 'TIMESTAMP' in query_params:
                        timestamp = query_params['TIMESTAMP'][0]
                        decoded_timestamp = urllib.parse.unquote(timestamp)
                        processed_row['Event Time'] = decoded_timestamp
                        
                except Exception as e:
                    logger.warning(f"处理Postback URL失败: {str(e)}")
                    # On decode failure, leave fields empty
                    processed_row['Attributed Touch Time'] = ''
                    processed_row['Install Time'] = ''
                    processed_row['Event Time'] = ''
            else:
                # If no Postback URL column, add empty time fields
                processed_row['Attributed Touch Time'] = ''
                processed_row['Install Time'] = ''
                processed_row['Event Time'] = ''
            
            processed_data.append(processed_row)
        
        return jsonify({
            'success': True,
            'data': processed_data,
            'processed_count': len(processed_data)
        })
        
    except Exception as e:
        logger.error(f"批量处理CSV Postback URL失败: {str(e)}")
        return jsonify({'error': f'处理失败：{str(e)}'}), 500

if __name__ == '__main__':
    # Limit Flask-SocketIO to one worker to avoid port conflicts
    socketio.run(app, host='0.0.0.0', port=5000, allow_unsafe_werkzeug=True, debug=False, use_reloader=False) 